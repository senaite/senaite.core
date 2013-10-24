from bika.lims import bikaMessageFactory as _
from bika.lims import logger
from bika.lims import PMF
from bika.lims.browser import BrowserView
from bika.lims.idserver import renameAfterCreation
from bika.lims.utils import changeWorkflowState
from bika.lims.utils import sortable_title
from bika.lims.utils import to_utf8 as _c
from cStringIO import StringIO
from DateTime import DateTime
from openpyxl.reader.excel import load_workbook
from os.path import join
from Persistence import PersistentMapping
from pkg_resources import resource_listdir, resource_filename, ResourceManager
from Products.Archetypes.config import REFERENCE_CATALOG
from Products.CMFCore.utils import getToolByName
from Products.CMFPlone.utils import safe_unicode
from Products.Five.browser.pagetemplatefile import ViewPageTemplateFile
from zipfile import ZipFile, ZIP_DEFLATED

import Globals
import json
import pkg_resources
import re
import tempfile
import transaction
import zope

try:
    from zope.component.hooks import getSite
except:
    # Plone < 4.3
    from zope.app.component.hooks import getSite


def validate(instance, skip=[]):
    errors = {}
    instance.Schema().validate(instance, instance.REQUEST, errors, True, True)
    for key in skip:
        if key in errors:
            del(errors[key])
    if errors:
        try:
            title = instance.Title()
        except:
            title = ''
        assert not errors, "Got errors validating {0} : {1}".format(
            instance, str(errors))


class LoadSetupData(BrowserView):

    def __init__(self, context, request):
        BrowserView.__init__(self, context, request)
        self.context = context
        self.request = request

        self.portal_catalog = getToolByName(self.context, 'portal_catalog')
        self.bc = bsc = getToolByName(self.context, 'bika_catalog')
        self.bsc = bsc = getToolByName(self.context, 'bika_setup_catalog')
        self.reference_catalog = getToolByName(self.context, REFERENCE_CATALOG)
        self.portal_registration = getToolByName(self.context, 'portal_registration')
        self.portal_groups = getToolByName(self.context, 'portal_groups')
        self.portal_membership = getToolByName(self.context, 'portal_membership')
        self.plone_utils = getToolByName(self.context, 'plone_utils')
        self.uc = getToolByName(context, 'uid_catalog')
        self.bac = getToolByName(context, 'bika_analysis_catalog')
        self.bsc = getToolByName(context, 'bika_setup_catalog')
        self.wf = getToolByName(context, 'portal_workflow')

        # dependencies to resolve
        self.deferred = []

        self.request.set('disable_border', 1)

    def set_wf_history(self, obj, hist):
        old_history = eval(hist)
        for wf_id,wfh in old_history.items():
            if obj.portal_type == "AnalysisRequest" and wf_id == "bika_analysis_workflow":
                wf_id = "bika_ar_workflow"
            if wf_id == 'bika_standardsample_workflow':
                wf_id = 'bika_referencesample_workflow'
            if wf_id == 'bika_standardanalysis_workflow':
                wf_id = 'bika_referenceanalysis_workflow'
            for state in wfh:
                if state == wfh[-1]:
                    set_permissions=False
                else:
                    set_permissions=True
                changeWorkflowState(obj, wf_id, state['review_state'],
                                    set_permissions=set_permissions,
                                    portal_workflow=self.wf, **state)

    def solve_deferred(self):
        # walk through self.deferred, linking ReferenceFields as we go
        unsolved = []
        for d in self.deferred:
            src_obj = d['src_obj']
            src_field = src_obj.getField(d['src_field'])
            multiValued = src_field.multiValued
            src_mutator = src_field.getMutator(src_obj)
            src_accessor = src_field.getAccessor(src_obj)

            tool = getToolByName(self.context, d['dest_catalog'])
            try:
                proxies = tool(d['dest_query'])
            except:
                continue
            if len(proxies) > 0:
                obj = proxies[0].getObject()
                if multiValued:
                    value = src_accessor()
                    value.append(obj.UID())
                else:
                    value = obj.UID()
                src_mutator(value)
            else:
                unsolved.append(d)
        self.deferred = unsolved
        return len(unsolved)

    def __call__(self):
        form = self.request.form

        portal = getSite()

        wb = None
        if 'setupexisting' in form and 'existing' in form and form['existing']:
                fn = form['existing']
                self.dataset_name = fn
                filename = resource_filename("bika.lims", "setupdata/%s/%s.xlsx" % (fn, fn))
                wb = load_workbook(filename = filename)
        elif 'setupfile' in form and 'file' in form and form['file']:
                tmp = tempfile.mktemp()
                file_content = form['file'].read()
                open(tmp, 'wb').write(file_content)
                wb = load_workbook(filename=tmp)
                self.dataset_name = 'uploaded'

        assert(wb != None)

        sheets = {}
        self.nr_rows = 0
        for sheetname in wb.get_sheet_names():
            if len(sheetname) > 31: print sheetname
            sheets[sheetname] = wb.get_sheet_by_name(sheetname)
            self.nr_rows += sheets[sheetname].get_highest_row()

        if 'Lab Information' in sheets:
            self.load_lab_information(sheets['Lab Information'])
        if 'Lab Contacts' in sheets:
            self.load_lab_contacts(sheets['Lab Contacts'])
        if 'Lab Departments' in sheets:
            self.load_lab_departments(sheets['Lab Departments'])
        if 'Clients' in sheets:
            self.load_clients(sheets['Clients'])
        if 'Client Contacts' in sheets:
            self.load_client_contacts(sheets['Client Contacts'])
        if 'Container Types' in sheets:
            self.load_containertypes(sheets["Container Types"])
        if 'Preservations' in sheets:
            self.load_preservations(sheets["Preservations"])
        if 'Containers' in sheets:
            self.load_containers(sheets["Containers"])
        if 'Suppliers' in sheets:
            self.load_suppliers(sheets['Suppliers'])
        if 'Supplier Contacts' in sheets:
            self.load_supplier_contacts(sheets['Supplier Contacts'])
        if 'Manufacturers' in sheets:
            self.load_manufacturers(sheets['Manufacturers'])
        if 'Instrument Types' in sheets:
            self.load_instrumenttypes(sheets['Instrument Types'])
        if 'Instruments' in sheets:
            self.load_instruments(sheets['Instruments'])
        if 'Instrument Validations' in sheets:
            self.load_instrumentvalidations(sheets['Instrument Validations'])
        if 'Instrument Calibrations' in sheets:
            self.load_instrumentcalibrations(sheets['Instrument Calibrations'])
        if 'Instrument Certifications' in sheets:
            self.load_instrumentcertifications(sheets['Instrument Certifications'])
        if 'Instrument Maintenance Tasks' in sheets:
            self.load_instrumentmaintenancetasks(sheets['Instrument Maintenance Tasks'])
        if 'Instrument Schedule' in sheets:
            self.load_instrumentschedule(sheets['Instrument Schedule'])
        if 'Sample Matrices' in sheets:
            self.load_sample_matrices(sheets['Sample Matrices'])

        if 'Batch Labels' in sheets:
            self.load_BatchLabels(sheets['Batch Labels'])

        if 'Sample Types' in sheets:
            self.load_sample_types(sheets['Sample Types'])
        if 'Sample Points' in sheets:
            self.load_sample_points(sheets['Sample Points'])
        if 'Sample Point Sample Types' in sheets:
            self.link_samplepoint_sampletype(sheets['Sample Point Sample Types'])
        if 'Analysis Categories' in sheets:
            self.load_analysis_categories(sheets['Analysis Categories'])
        if 'Methods' in sheets:
            self.load_methods(sheets['Methods'])
        if 'Calculation Interim Fields' in sheets:
            self.load_interim_fields(sheets['Calculation Interim Fields'])
        if 'AnalysisService InterimFields' in sheets:
            self.load_service_interims(sheets['AnalysisService InterimFields'])
        #if 'Lab Products' in sheets:
        #    self.load_lab_products(sheets['Lab Products'])
        if 'Sampling Deviations' in sheets:
            self.load_sampling_deviations(sheets['Sampling Deviations'])
        if 'Reference Manufacturers' in sheets:
            self.load_reference_manufacturers(sheets['Reference Manufacturers'])
        if 'Calculations' in sheets:
            self.load_calculations(sheets['Calculations'])
        if 'Analysis Services' in sheets:
            self.load_analysis_services(sheets['Analysis Services'])
        if 'AnalysisService ResultOptions' in sheets:
            self.service_result_options(sheets['AnalysisService ResultOptions'])
        if 'Analysis Service Uncertainties' in sheets:
            self.service_uncertainties(sheets['Analysis Service Uncertainties'])

        if 'Analysis Specifications' in sheets:
            self.load_analysis_specifications(sheets['Analysis Specifications'])
        if 'Analysis Profile Services' in sheets:
            self.load_analysis_profile_services(sheets['Analysis Profile Services'])
        if 'Analysis Profiles' in sheets:
            self.load_analysis_profiles(sheets['Analysis Profiles'])
        if 'AR Template Analyses' in sheets:
            self.load_artemplate_analyses(sheets['AR Template Analyses'])
        if 'AR Template Partitions' in sheets:
            self.load_artemplate_partitions(sheets['AR Template Partitions'])
        if 'AR Templates' in sheets:
            self.load_artemplates(sheets['AR Templates'])
        if 'Reference Definition Results' in sheets:
            self.load_reference_definition_results(sheets['Reference Definition Results'])
        if 'Reference Definitions' in sheets:
            self.load_reference_definitions(sheets['Reference Definitions'])
        if 'Reference Suppliers' in sheets:
            self.load_reference_suppliers(sheets['Reference Suppliers'])
        if 'Reference Supplier Contacts' in sheets:
            self.load_reference_supplier_contacts(sheets['Reference Supplier Contacts'])

        if 'Worksheet Template Layouts' in sheets:
            self.load_wst_layouts(sheets['Worksheet Template Layouts'])
        if 'Worksheet Template Services' in sheets:
            self.load_wst_services(sheets['Worksheet Template Services'])
        if 'Worksheet Templates' in sheets:
            self.load_worksheet_templates(sheets['Worksheet Templates'])

#        if 'Reference Sample Results' in sheets:
#            self.load_reference_sample_results(sheets['Reference Sample Results'])
#        if 'Reference Samples' in sheets:
#            self.load_reference_samples(sheets['Reference Samples'])
#        if 'Reference Analyses Interims' in sheets:
#            self.load_reference_analyses_interims(sheets['Reference Analyses Interims'])
#        if 'Reference Analyses' in sheets:
#            self.load_reference_analyses(sheets['Reference Analyses'])
#        if 'Samples' in sheets:
#            self.load_samples(sheets['Samples'])
#        if 'AR CCContacts' in sheets:
#            self.load_arccs(sheets['AR CCContacts'])
#        if 'Analysis Requests' in sheets:
#            self.load_ars(sheets['Analysis Requests'])
#        if 'Analyses Interims' in sheets:
#            self.load_analyses_interims(sheets['Analyses Interims'])
#        if 'Analyses' in sheets:
#            self.load_analyses(sheets['Analyses'])
#        if 'Analysis Requests' in sheets:
#            for arid, wfh in self.ar_workflows.items():
#                self.set_wf_history(self.ars[arid], wfh)
#        if 'Worksheets' in sheets:
#            self.load_worksheets(sheets['Worksheets'])
#        if 'Duplicate Analyses Interims' in sheets:
#            self.load_duplicate_interims(sheets['Duplicate Analyses Interims'])
#        if 'Duplicate Analyses' in sheets:
#            self.load_duplicate_analyses(sheets['Duplicate Analyses'])
#        if 'Worksheet Layouts' in sheets:
#            self.load_worksheet_layouts(sheets['Worksheet Layouts'])

        if 'Setup' in sheets:
            self.load_bika_setup(sheets['Setup'])
        if 'ID Prefixes' in sheets:
            self.load_id_prefixes(sheets['ID Prefixes'])
        if 'Attachment Types' in sheets:
            self.load_attachment_types(sheets['Attachment Types'])

        check = len(self.deferred)
        while len(self.deferred) > 0:
            new = self.solve_deferred()
            logger.info("solved %s of %s deferred references" % (check - new, check))
            if new == check:
                raise Exception("%s unsolved deferred references: %s"%(len(self.deferred), self.deferred))
            check = new

        logger.info("Rebuilding bika_setup_catalog")
        bsc = getToolByName(self.context, 'bika_setup_catalog')
        bsc.clearFindAndRebuild()
        logger.info("Rebuilding bika_catalog")
        bc = getToolByName(self.context, 'bika_catalog')
        bc.clearFindAndRebuild()
        logger.info("Rebuilding bika_analysis_catalog")
        bac = getToolByName(self.context, 'bika_analysis_catalog')
        bac.clearFindAndRebuild()


        message = PMF("Changes saved.")
        self.plone_utils.addPortalMessage(message)
        self.request.RESPONSE.redirect(portal.absolute_url())

    def load_bika_setup(self,sheet):
        logger.info("Loading Bika Setup...")
        values = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            values[row['Field']] = row['Value']

        DSL = {'days': int(values['DefaultSampleLifetime_days'] and values['DefaultSampleLifetime_days'] or 0),
               'hours': int(values['DefaultSampleLifetime_hours'] and values['DefaultSampleLifetime_hours'] or 0),
               'minutes': int(values['DefaultSampleLifetime_minutes'] and values['DefaultSampleLifetime_minutes'] or 0),
               }
        self.context.bika_setup.edit(
            PasswordLifetime = int(values['PasswordLifetime']),
            AutoLogOff = int(values['AutoLogOff']),
            Currency = values['Currency'],
            MemberDiscount = str(float(values['MemberDiscount'])),
            VAT = str(float(values['VAT'])),
            MinimumResults = int(values['MinimumResults']),
            BatchEmail = int(values['BatchEmail']),
##            BatchFax = int(values['BatchFax']),
##            SMSGatewayAddress = values['SMSGatewayAddress'],
            SamplingWorkflowEnabled = values['SamplingWorkflowEnabled'],
            CategoriseAnalysisServices = self.to_bool(values['CategoriseAnalysisServices']),
            EnableAnalysisRemarks = self.to_bool(values.get('EnableAnalysisRemarks', '')),
            DryMatterService = self.services.get(values['DryMatterService'], ''),
            ARImportOption = values['ARImportOption'],
            ARAttachmentOption = values['ARAttachmentOption'][0].lower(),
            AnalysisAttachmentOption = values['AnalysisAttachmentOption'][0].lower(),
            DefaultSampleLifetime = DSL,
            AutoPrintLabels = values['AutoPrintLabels'].lower(),
            AutoLabelSize = values['AutoLabelSize'].lower(),
            YearInPrefix = self.to_bool(values['YearInPrefix']),
            SampleIDPadding = int(values['SampleIDPadding']),
            ARIDPadding = int(values['ARIDPadding']),
            ExternalIDServer = self.to_bool(values['ExternalIDServer']),
            IDServerURL = values['IDServerURL'],
        )

    def load_id_prefixes(self,sheet):
        logger.info("Loading ID Prefixes...")
        prefixes = self.context.bika_setup.getPrefixes()
        rows = self.get_rows(sheet, 3)
        for row in rows:
            # remove existing prefix from list
            prefixes = [p for p in prefixes
                        if p['portal_type'] != row['portal_type']]
            # add new prefix to list
            prefixes.append({'portal_type': row['portal_type'],
                             'padding': row['padding'],
                             'prefix': row['prefix']})
        self.context.bika_setup.setPrefixes(prefixes)

    def load_containertypes(self, sheet):
        logger.info("Loading Container Types...")
        folder = self.context.bika_setup.bika_containertypes
        self.containertypes = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('ContainerType', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.containertypes[row['title']] = obj

    def load_BatchLabels(self, sheet):
        logger.info("Loading Batch Labels...")
        folder = self.context.bika_setup.bika_batchlabels
        self.batchlabels = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('BatchLabel', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'])
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.batchlabels[row['title']] = obj

    def load_CaseStatuses(self, sheet):
        logger.info("Loading Batch Labels...")
        folder = self.context.bika_setup.bika_casestatuses
        self.casestatuses = {}
        rows = self.get_rows(sheet, 3)
        for row in rows[3:]:
            if row['title']:
                _id = folder.invokeFactory('CaseStatus', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.casestatuses[row['title']] = obj

    def load_CaseOutcomes(self, sheet):
        logger.info("Loading Case Outcomes...")
        folder = self.context.bika_setup.bika_caseoutcomes
        self.caseoutcomes = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('CaseOutcome', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                # renameAfterCreation(obj)
                self.caseoutcomes[row['title']] = obj

    def load_preservations(self, sheet):
        logger.info("Loading Preservations...")
        folder = self.context.bika_setup.bika_preservations
        self.preservations = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('Preservation', id = 'tmp')
                obj = folder[_id]
                RP = {'days': int(row['RetentionPeriod_days'] and row['RetentionPeriod_days'] or 0),
                      'hours': int(row['RetentionPeriod_hours'] and row['RetentionPeriod_hours'] or 0),
                      'minutes': int(row['RetentionPeriod_minutes'] and row['RetentionPeriod_minutes'] or 0),
                      }

                obj.edit(title = row['title'],
                         description = row.get('description', ''),
                         RetentionPeriod = RP)
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.preservations[row['title']] = obj

    def load_containers(self, sheet):
        logger.info("Loading Containers...")
        folder = self.context.bika_setup.bika_containers
        self.containers = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('Container', id = 'tmp')
                obj = folder[_id]
                P = row['Preservation_title']
                obj.edit(title = row['title'],
                         description = row.get('description', ''),
                         Capacity = row.get('Capacity',0),
                         PrePreserved = self.to_bool(row['PrePreserved']))
                if row['ContainerType_title']:
                    obj.setContainerType(self.containertypes[row['ContainerType_title']])
                if P:
                    obj.setPreservation(self.preservations[P])
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.containers[row['title']] = obj

    def load_lab_information(self, sheet):
        logger.info("Loading Lab information...")
        laboratory = self.context.bika_setup.laboratory
        values = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            values[row['Field']] = row['Value']

        addresses = {}
        for add_type in ['Physical', 'Postal', 'Billing']:
            addresses[add_type] = {}
            for key in ['Address', 'City', 'State', 'Zip', 'Country']:
                addresses[add_type][key] = str(values["%s_%s" % (add_type, key)])

        if values['AccreditationBodyLogo']:
            path = resource_filename("bika.lims","setupdata/%s/%s" \
                                     % (self.dataset_name, values['AccreditationBodyLogo']))
            file_data = open(path, "rb").read()
        else:
            file_data = None

        laboratory.edit(
            Name = values['Name'],
            LabURL = values['LabURL'],
            Confidence = values['Confidence'],
            LaboratoryAccredited = self.to_bool(values['LaboratoryAccredited']),
            AccreditationBodyLong = values['AccreditationBodyLong'],
            AccreditationBody = values['AccreditationBody'],
            AccreditationBodyURL = values['AccreditationBodyURL'],
            Accreditation = values['Accreditation'],
            AccreditationReference = values['AccreditationReference'],
            AccreditationBodyLogo = file_data,
            TaxNumber = values['TaxNumber'],
            Phone = values['Phone'],
            Fax = values['Fax'],
            EmailAddress = values['EmailAddress'],
            PhysicalAddress = addresses['Physical'],
            PostalAddress = addresses['Postal'],
            BillingAddress = addresses['Billing']
        )

    def load_lab_contacts(self, sheet):
        logger.info("Loading Lab contacts...")
        self.lab_contacts = {}
        folder = self.context.bika_setup.bika_labcontacts
        rows = self.get_rows(sheet, 3)
        for row in rows:

            ## Create LabContact

            if row['Signature']:
                path = resource_filename("bika.lims","setupdata/%s/%s" \
                                         % (self.dataset_name, row['Signature']))
                file_data = open(path, "rb").read()
            else:
                file_data = None

            if not row['Firstname']:
                continue

            _id = folder.invokeFactory('LabContact', id='tmp')
            obj = folder[_id]
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)
            Fullname = row['Firstname'] + " " + row.get('Surname','')
            obj.edit(
                title = Fullname,
                Salutation = row.get('Salutation', ''),
                Firstname = row['Firstname'],
                Surname = row.get('Surname',''),
                EmailAddress = row.get('EmailAddress',''),
                BusinessPhone = row.get('BusinessPhone',''),
                BusinessFax = row.get('BusinessFax',''),
                HomePhone = row.get('HomePhone',''),
                MobilePhone = row.get('MobilePhone',''),
                JobTitle = row.get('JobTitle',''),
                Username = row.get('Username',''),
                Signature = file_data
            )
            validate(obj)
            self.lab_contacts[Fullname] = obj

            if row['Department_title']:
                self.deferred.append({'src_obj': obj,
                                      'src_field': 'Department',
                                      'dest_catalog': 'bika_setup_catalog',
                                      'dest_query': {'portal_type':'Department',
                                                     'title':row['Department_title']}})

            ## Create Plone user
            username = safe_unicode(row['Username']).encode('utf-8')
            password = safe_unicode(row['Password']).encode('utf-8')
            if(username):
                member = self.portal_registration.addMember(
                    username,
                    password,
                    properties = {
                        'username': username,
                        'email': row['EmailAddress'],
                        'fullname': Fullname}
                    )
                group_ids = [g.strip() for g in row['Groups'].split(',')]
                role_ids = [r.strip() for r in row['Roles'].split(',')]
                # Add user to all specified groups
                for group_id in group_ids:
                    group = self.portal_groups.getGroupById(group_id)
                    if group:
                        group.addMember(username)
                # Add user to all specified roles
                for role_id in role_ids:
                    member._addRole(role_id)
                # If user is in LabManagers, add Owner local role on clients folder
                if 'LabManager' in group_ids:
                    self.context.clients.manage_setLocalRoles(username, ['Owner',] )

    def load_lab_departments(self, sheet):
        logger.info("Loading Lab departments...")
        self.departments = {}
        lab_contacts = [o.getObject() for o in self.bsc(portal_type="LabContact")]
        folder = self.context.bika_setup.bika_departments
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('Department', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                manager = None
                for contact in lab_contacts:
                    if contact.getUsername() == row['LabContact_Username']:
                        manager = contact
                        break
                else:
                    message = "Department: lookup of '%s' in LabContacts/Username failed."%row['LabContact_Username']
                    logger.info(message)
                if manager:
                    obj.setManager(manager.UID())
                self.departments[row['title']] = obj
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_clients(self, sheet):
        logger.info("Loading Clients...")
        self.clients = {}
        folder = self.context.clients
        rows = self.get_rows(sheet, 3)
        for row in rows:
            addresses = {}
            for add_type in ['Physical', 'Postal', 'Billing']:
                addresses[add_type] = {}
                for key in ['Address', 'City', 'State', 'Zip', 'Country']:
                    addresses[add_type][key] = str(row["%s_%s" % (add_type, key)])

            _id = folder.invokeFactory('Client', id = 'tmp')
            obj = folder[_id]
            if not row['Name']:
                message = "Client %s has no Name"
                raise Exception(message)
            if not row['ClientID']:
                message = "Client %s has no Client ID"
                raise Exception(message)
            obj.edit(Name = row['Name'],
                     ClientID = row['ClientID'],
                     MemberDiscountApplies = row['MemberDiscountApplies'] and True or False,
                     BulkDiscount = row['BulkDiscount'] and True or False,
                     TaxNumber = row.get('TaxNumber',''),
                     Phone = row.get('Phone',''),
                     Fax = row.get('Fax',''),
                     EmailAddress = row.get('EmailAddress',''),
                     PhysicalAddress = addresses['Physical'],
                     PostalAddress = addresses['Postal'],
                     BillingAddress = addresses['Billing'],
                     AccountNumber = row.get('AccountNumber','')
            )
            obj.unmarkCreationFlag()
            validate(obj)
            renameAfterCreation(obj)
            self.clients[row['Name']] = obj

    def load_client_contacts(self, sheet):
        logger.info("Loading Client contacts...")
        folder = self.context.clients
        self.client_contacts = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            client = self.portal_catalog(portal_type = "Client",
                                         Title = row['Client_title'])
            if len(client) == 0:
                raise IndexError("Client invalid: '%s'" % row['Client_title'])
            client = client[0].getObject()
            _id = client.invokeFactory('Contact', id = 'tmp')
            contact = client[_id]
            fullname = "%(Firstname)s %(Surname)s" % row
            addresses = {}
            for add_type in ['Physical', 'Postal']:
                addresses[add_type] = {}
                for key in ['Address', 'City', 'State', 'Zip', 'Country']:
                    addresses[add_type][key] = str(row["%s_%s" % (add_type, key)])
            contact.edit(Salutation = row.get('Salutation',''),
                         Firstname = row.get('Firstname',''),
                         Surname = row.get('Surname',''),
                         Username = row.get('Username',''),
                         JobTitle = row.get('JobTitle',''),
                         Department = row.get('Department',''),
                         BusinessPhone = row.get('BusinessPhone',''),
                         BusinessFax = row.get('BusinessFax',''),
                         HomePhone = row.get('HomePhone',''),
                         MobilePhone = row.get('MobilePhone',''),
                         EmailAddress = row.get('EmailAddress',''),
                         PublicationPreference =  row.get('PublicationPreference','').split(","),
                         AttachmentsPermitted = row['AttachmentsPermitted'] and True or False,
                         PhysicalAddress = addresses['Physical'],
                         PostalAddress = addresses['Postal'],
            )

            contact.unmarkCreationFlag()
            validate(contact)
            renameAfterCreation(contact)
            self.client_contacts[fullname] = contact

            # CC Contacts
            if row['CCContacts']:
                for _fullname in row['CCContacts'].split(","):
                    self.deferred.append({'src_obj': contact,
                                          'src_field': 'CCContact',
                                          'dest_catalog': 'portal_catalog',
                                          'dest_query': {'portal_type':'Contact',
                                                         'getFullname':_fullname}})

            ## Create Plone user
            username = safe_unicode(row['Username']).encode('utf-8')
            password = safe_unicode(row['Password']).encode('utf-8')
            if(row['Username']):
                try:
                    member = self.portal_registration.addMember(
                        username,
                        row['Password'],
                        properties = {
                            'username': username,
                            'email': row['EmailAddress'],
                            'fullname': fullname}
                        )
                except:
                    logger.info("Error adding user (already exists?): %s" % username)
                contact.aq_parent.manage_setLocalRoles(username, ['Owner',] )
                # add user to Clients group
                group = self.portal_groups.getGroupById('Clients')
                group.addMember(username)

    def load_instruments(self, sheet):
        logger.info("Loading Instruments...")
        folder = self.context.bika_setup.bika_instruments
        self.instruments = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if (not row['Type']
                or not row['title']
                or not row['Supplier']
                or not row['Brand']):
                continue

            _id = folder.invokeFactory('Instrument', id='tmp')
            obj = folder[_id]

            obj.edit(title=row['title'],
                     description=row.get('description', ''),
                     Type=row['Type'],
                     Brand=row['Brand'],
                     Model=row['Model'],
                     SerialNo=row.get('SerialNo', ''),
                     CalibrationCertificate=row.get('CalibrationCertificate', ''),
                     CalibrationExpiryDate=row.get('CalibrationExpiryDate', ''),
                     DataInterface=row.get('DataInterface', ''))

            obj.setInstrumentType(self.instrumenttypes[row['Type']])
            obj.setManufacturer(self.manufacturers[row['Brand']])
            obj.setSupplier(self.suppliers[row['Supplier']])
            self.instruments[row['title']] = obj.UID()
            obj.unmarkCreationFlag()
            validate(obj)
            renameAfterCreation(obj)

    def load_sample_points(self, sheet):
        logger.info("Loading Sample points...")
        setup_folder = self.context.bika_setup.bika_samplepoints
        self.samplepoints = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['title']:
                continue
            if row['Client_title']:
                client_title = row['Client_title']
                client = self.portal_catalog(portal_type = "Client",
                                             Title = client_title)
                if len(client) == 0:
                    raise IndexError("Sample Point %s: Client invalid: '%s'" %\
                                     (row['title'], client_title))
                folder = client[0].getObject()
            else:
                folder = setup_folder

            if row['Latitude']: logger.log("Ignored SamplePoint Latitude", 'error')
            if row['Longitude']: logger.log("Ignored SamplePoint Longitude", 'error')

            _id = folder.invokeFactory('SamplePoint', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = row['title'],
                     description = row.get('description', ''),
                     Composite = self.to_bool(row['Composite']),
                     Elevation = row['Elevation'],
                     SampleTypes = row['SampleType_title'] and self.sampletypes[row['SampleType_title']] or []
                     )
            self.samplepoints[row['title']] = obj
            obj.unmarkCreationFlag()
            validate(obj)
            renameAfterCreation(obj)

    def load_sample_types(self, sheet):
        logger.info("Loading Sample types...")
        folder = self.context.bika_setup.bika_sampletypes
        self.sampletypes = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('SampleType', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''),
                         RetentionPeriod = {'days':row['RetentionPeriod'] and row['RetentionPeriod'] or 0,'hours':0,'minutes':0},
                         Hazardous = self.to_bool(row['Hazardous']),
                         SampleMatrix = row['SampleMatrix_title'] and self.samplematrices[row['SampleMatrix_title']] or None,
                         Prefix = row['Prefix'],
                         MinimumVolume = row['MinimumVolume'],
                         ContainerType = row['ContainerType_title'] and self.containertypes[row['ContainerType_title']] or None
                )
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.sampletypes[row['title']] = obj

    def link_samplepoint_sampletype(self, sheet):
        logger.info("Loading Sample point - sample types...")
        rows = self.get_rows(sheet, 3)
        for row in rows:

            st = self.sampletypes[row['SampleType_title']]
            sp = self.samplepoints[row['SamplePoint_title']]

            sampletypes = sp.getSampleTypes()
            if st not in sampletypes:
                sampletypes.append(st)
                sp.setSampleTypes(sampletypes)

            samplepoints = st.getSamplePoints()
            if sp not in samplepoints:
                samplepoints.append(sp)
                st.setSamplePoints(samplepoints)

    def load_sampling_deviations(self, sheet):
        logger.info("Loading Sample deviations...")
        folder = self.context.bika_setup.bika_samplingdeviations
        self.samplingdeviations = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('SamplingDeviation', id = 'tmp')
                obj = folder[_id]
                self.samplingdeviations[row['title']] = obj.UID()
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_sample_matrices(self, sheet):
        logger.info("Loading Sample matrices...")
        folder = self.context.bika_setup.bika_samplematrices
        self.samplematrices = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('SampleMatrix', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.samplematrices[row['title']] = obj

    def load_analysis_categories(self, sheet):
        logger.info("Loading Analysis categories...")
        folder = self.context.bika_setup.bika_analysiscategories
        self.cats = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('AnalysisCategory', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                if row['Department_title']:
                    obj.setDepartment(self.departments[row['Department_title']].UID())
                self.cats[row['title']] = obj
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_methods(self, sheet):
        logger.info("Loading Methods...")
        folder = self.context.methods
        self.methods = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('Method', id = 'tmp')
                obj = folder[_id]

                obj.edit(title = row['title'],
                         description = row.get('description', ''),
                         Instructions = row.get('Instructions',''))

                if row['MethodDocument']:
                    path = resource_filename("bika.lims",
                                             "setupdata/%s/%s" \
                                             % (self.dataset_name, row['MethodDocument']))
                    #file_id = obj.invokeFactory("File", id=row['MethodDocument'])
                    #thisfile = obj[file_id]
                    file_data = open(path, "rb").read()
                    obj.setMethodDocument(file_data)

                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.methods[row['title']] = obj

    def load_service_interims(self, sheet):
        logger.info("Loading Analysis Service Interim fields...")
        # Read all InterimFields into self.service_interims
        self.service_interims = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            service_title = row['Service_title']
            if service_title not in self.service_interims.keys():
                self.service_interims[service_title] = []
            self.service_interims[service_title].append({
                'keyword': row['keyword'],
                'title': row.get('title', ''),
                'type': 'int',
                'value': row['value'],
                'unit': row['unit'] and row['unit'] or ''})

    def load_analysis_services(self, sheet):
        logger.info("Loading Analysis Services...")
        folder = self.context.bika_setup.bika_analysisservices
        self.services = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['title']:
                continue

            _id = folder.invokeFactory('AnalysisService', id = 'tmp')
            obj = folder[_id]
            MTA = {'days': int(row['MaxTimeAllowed_days'] and row['MaxTimeAllowed_days'] or 0),
                   'hours': int(row['MaxTimeAllowed_days'] and row['MaxTimeAllowed_days'] or 0),
                   'minutes': int(row['MaxTimeAllowed_minutes'] and row['MaxTimeAllowed_minutes'] or 0),
                   }
            obj.edit(
                title = row['title'],
                description = row.get('description', ''),
                Keyword = row['Keyword'],
                PointOfCapture = row['PointOfCapture'],
                Category = self.cats[row['AnalysisCategory_title']].UID(),
                Department = row['Department_title'] and self.departments[row['Department_title']].UID() or None,
                ReportDryMatter = self.to_bool(row['ReportDryMatter']),
                AttachmentOption = row['Attachment'][0].lower(),
                Unit = row['Unit'] and row['Unit'] or None,
                Precision = row['Precision'] and str(row['Precision']) or '2',
                MaxTimeAllowed = MTA,
                Price = row['Price'] and "%02f"%(float(row['Price'])) or "0,00",
                BulkPrice = row['BulkPrice'] and "%02f"%(float(row['BulkPrice'])) or "0.00",
                VAT = row['VAT'] and "%02f"%(float(row['VAT'])) or "0.00",
                Method = row['Method'] and self.methods[row['Method']] or None,
                Instrument = row['Instrument_title'] and self.instruments[row['Instrument_title']] or None,
                Calculation = row['Calculation_title'] and self.calcs[row['Calculation_title']] or None,
                DuplicateVariation = "%02f" % float(row['DuplicateVariation']),
                Accredited = self.to_bool(row['Accredited']),
                InterimFields = hasattr(self,'service_interims') and self.service_interims.get(row['title'], []) or []
            )
            self.services[row['title']] = obj
            obj.unmarkCreationFlag()
            validate(obj)
            renameAfterCreation(obj)

    def service_result_options(self, sheet):
        logger.info("Loading Analysis Service result options...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            service = self.services[row['Service_title']]
            sro = service.getResultOptions()
            sro.append({'ResultValue':row['ResultValue'],
                        'ResultText':row['ResultText']})
            service.setResultOptions(sro)

    def service_uncertainties(self, sheet):
        logger.info("Loading Analysis Service uncertainities...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            service = self.services[row['Service_title']]
            sru = service.getUncertainties()
            sru.append({'intercept_min':row['Range Min'],
                        'intercept_max':row['Range Max'],
                        'errorvalue': row['Uncertainty Value']})
            service.setUncertainties(sru)

    def load_interim_fields(self, sheet):
        logger.info("Loading Calculations Interim fields...")
        # Read all InterimFields into self.interim_fields
        self.interim_fields = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            calc_title = row['Calculation_title']
            if calc_title not in self.interim_fields.keys():
                self.interim_fields[calc_title] = []
            self.interim_fields[calc_title].append({
                'keyword': row['keyword'],
                'title': row.get('title', ''),
                'type': 'int',
                'hidden': ('hidden' in row and row['hidden']) and True or False,
                'value': row['value'],
                'unit': row['unit'] and row['unit'] or ''})

    def load_calculations(self, sheet):
        logger.info("Loading Calculations...")
        self.calcs = {}
        folder = self.context.bika_setup.bika_calculations
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['title']:
                continue
            calc_title = row['title']
            calc_interims = self.interim_fields.get(calc_title, [])
            formula = row['Formula']
            # scan formula for dep services
            keywords = re.compile(r"\[([^\]]+)\]").findall(formula)
            # remove interims from deps
            interim_keys = [k['keyword'] for k in calc_interims]
            dep_keywords = [k for k in keywords if k not in interim_keys]

            _id = folder.invokeFactory('Calculation', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = calc_title,
                     description = row.get('description', ''),
                     InterimFields = calc_interims,
                     Formula = str(row['Formula']))
            for kw in dep_keywords:
                # If the keyword is in calc_interims, no service dependency gets deferred
                self.deferred.append({'src_obj': obj,
                                      'src_field': 'DependentServices',
                                      'dest_catalog': 'bika_setup_catalog',
                                      'dest_query': {'portal_type':'AnalysisService',
                                                     'getKeyword':kw}})
            self.calcs[row['title']] = obj
            obj.unmarkCreationFlag()
            validate(obj)
            renameAfterCreation(obj)

    def load_analysis_profile_services(self, sheet):
        logger.info("Loading Analysis Profile Services...")
        # Read the Analysis Profile Services into self.profile_services
        self.profile_services = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['Profile'] not in self.profile_services.keys():
                self.profile_services[row['Profile']] = []
            self.profile_services[row['Profile']].append(
                self.services[row['Service']])

    def load_analysis_profiles(self, sheet):
        logger.info("Loading Analysis Profiles...")
        folder = self.context.bika_setup.bika_analysisprofiles
        # Read the Analysis Profile Services into a dict
        self.profiles = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('AnalysisProfile', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''),
                         ProfileKey = row['ProfileKey'])
                obj.setService(self.profile_services[row['title']])
                obj.unmarkCreationFlag()
                validate(obj, skip=['Service',])
                renameAfterCreation(obj)
                self.profiles[row['title']] = obj

    def load_artemplate_analyses(self, sheet):
        logger.info("Loading Analysis Request Template analyses...")
        # Read the AR Template Services into self.artemplate_analyses
        self.artemplate_analyses = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['ARTemplate'] not in self.artemplate_analyses.keys():
                self.artemplate_analyses[row['ARTemplate']] = []
            self.artemplate_analyses[row['ARTemplate']].append({
                'service_uid': self.services[row['service_uid']].UID(),
                'partition': row['partition']})

    def load_artemplate_partitions(self, sheet):
        logger.info("Loading Analysis Request Template partitions...")
        # Read the AR Template Partitions into self.artemplate_partitions
        self.artemplate_partitions = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['ARTemplate'] not in self.artemplate_partitions.keys():
                self.artemplate_partitions[row['ARTemplate']] = []
            self.artemplate_partitions[row['ARTemplate']].append({
                'part_id': row['part_id'],
                'container_uid': row['container']
                and self.containers[row['container']].UID() or [],
                'preservation_uid': row['preservation']
                and self.preservations[row['preservation']].UID() or row['preservation']})

    def load_artemplates(self, sheet):
        logger.info("Loading Analysis Request Templates...")
        self.artemplates = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['title']:
                continue
            analyses = self.artemplate_analyses[row['title']]
            client_title = row['Client_title'] or 'lab'
            if row['title'] in self.artemplate_partitions:
                partitions = self.artemplate_partitions[row['title']]
            else:
                partitions = [{'part_id': 'part-1',
                               'container': '',
                               'preservation': ''}]

            if client_title == 'lab':
                folder = self.context.bika_setup.bika_artemplates
            else:
                folder = self.clients[client_title]

            if row['SampleType_title']:
                sampletypes = row['SampleType_title']
            else:
                sampletypes = None
            if row['SamplePoint_title']:
                samplepoints = row['SamplePoint_title']
            else:
                samplepoints = None

            _id = folder.invokeFactory('ARTemplate', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = row['title'],
                     description = row.get('description', ''),
                     Remarks = row.get('Remarks',''),
                     ReportDryMatter = bool(row['ReportDryMatter']))
            obj.setSampleType(sampletypes)
            obj.setSamplePoint(samplepoints)
            obj.setPartitions(partitions)
            obj.setAnalyses(analyses)
            obj.unmarkCreationFlag()
            validate(obj)
            renameAfterCreation(obj)
            self.artemplates[row['title']] = obj.UID()

    def load_reference_definition_results(self, sheet):
        logger.info("Loading Reference definition results...")
        self.ref_def_results = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['ReferenceDefinition_title'] \
               not in self.ref_def_results.keys():
                self.ref_def_results[
                    row['ReferenceDefinition_title']] = []
            self.ref_def_results[
                row['ReferenceDefinition_title']].append({
                    'uid': self.services[row['service']].UID(),
                    'result': row['result'],
                    'min': row['min'],
                    'max': row['max']})

    def load_reference_definitions(self, sheet):
        logger.info("Loading Reference definitions...")
        self.definitions = {}
        folder = self.context.bika_setup.bika_referencedefinitions
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('ReferenceDefinition', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''),
                         Blank = self.to_bool(row['Blank']),
                         ReferenceResults = self.ref_def_results.get(row['title'], []),
                         Hazardous = self.to_bool(row['Hazardous']))
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.definitions[row['title']] = obj.UID()

    def load_analysis_specifications(self, sheet):
        logger.info("Loading Analysis Specifications...")
        # First we sort the specs by client/sampletype
        #  { Client: { SampleType: { service, min, max, error }... }... }
        all_specs = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            client_title = row['Client_title'] or 'lab'
            sampletype_title = row['SampleType_title']
            if client_title not in all_specs:
                all_specs[client_title] = {}
            if sampletype_title not in all_specs[client_title]:
                all_specs[client_title][sampletype_title] = []
            all_specs[client_title][sampletype_title].append({
                'keyword': str(self.services[row['service']].getKeyword()),
                'min': str(row['min']),
                'max': str(row['max']),
                'error': str(row['error'])})
        for client, client_specs in all_specs.items():
            if client == 'lab':
                folder = self.context.bika_setup.bika_analysisspecs
            else:
                folder = self.clients[client]
            for sampletype_title, resultsrange in client_specs.items():
                sampletype = self.sampletypes[sampletype_title]
                _id = folder.invokeFactory('AnalysisSpec', id = 'tmp')
                obj = folder[_id]
                obj.edit(
                         title = sampletype.Title(),
                         ResultsRange = resultsrange)
                obj.setSampleType(sampletype.UID())
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_reference_manufacturers(self, sheet):
        logger.info("Loading Reference Manufacturers...")
        folder = self.context.bika_setup.bika_manufacturers
        self.ref_manufacturers = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('Manufacturer', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                self.ref_manufacturers[row['title']] = obj.UID()
                renameAfterCreation(obj)

    def load_reference_suppliers(self, sheet):
        logger.info("Loading Reference Suppliers...")
        folder = self.context.bika_setup.bika_suppliers
        self.ref_suppliers = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            _id = folder.invokeFactory('Supplier', id = 'tmp')
            obj = folder[_id]
            obj.edit(AccountNumber = row['AccountNumber'],
                     Name = row['Name'],
                     EmailAddress = row['EmailAddress'],
                     Phone = row['Phone'],
                     Fax = row['Fax'])
            obj.unmarkCreationFlag()
            validate(obj)
            self.ref_suppliers[obj.Title()] = obj
            renameAfterCreation(obj)

    def load_reference_supplier_contacts(self, sheet):
        logger.info("Loading Reference Supplier Contacts...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['ReferenceSupplier_Name']:
                continue
            folder = self.bsc(portal_type="Supplier",
                              Title = row['ReferenceSupplier_Name'])
            if (len(folder) > 0):
                folder = folder[0].getObject()
                _id = folder.invokeFactory('SupplierContact', id = 'tmp')
                obj = folder[_id]
                obj.edit(
                    Firstname = row['Firstname'],
                    Surname = row['Surname'],
                    EmailAddress = row['EmailAddress'])
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

                if 'Username' in row:
    ##               'Password' in row:
    ##                self.context.REQUEST.set('username', _c(row['Username']))
    ##                self.context.REQUEST.set('password', _c(row['Password']))
    ##                self.context.REQUEST.set('email', _c(row['EmailAddress']))
    ##                pr = getToolByName(self.context, 'portal_registration')
    ##                pr.addMember(_c(row['Username']),
    ##                             _c(row['Password']),
    ##                             properties = {
    ##                                 'username': _c(row['Username']),
    ##                                 'email': _c(row['EmailAddress']),
    ##                                 'fullname': " ".join((row['Firstname'],
    ##                                                       row['Surname']))})
                    obj.setUsername(row['Username'])

    def load_reference_sample_results(self, sheet):
        logger.info("Loading Reference Supplier Contacts...")
        # Read the Ref Sample Results into self.refsample_results
        self.refsample_results = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            refsample_id = row['ReferenceSample_id']
            if not self.refsample_results.has_key(refsample_id):
                self.refsample_results[refsample_id] = []
            self.refsample_results[refsample_id].append({
                'uid': self.services[row['service']].UID(),
                'result': row['result'],
                'min': row['min'],
                'max': row['max']})

    def load_reference_samples(self, sheet):
        logger.info("Loading Reference Samples...")
        self.ref_samples = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            supplier = self.ref_suppliers[row['ReferenceSupplier_title']]
            supplier.invokeFactory('ReferenceSample', id = row['id'])
            obj = supplier[row['id']]
            ref_def = row['ReferenceDefinition_title']
            ref_def = ref_def and self.definitions[ref_def] or ''
            ref_man = row['ReferenceManufacturer_title']
            ref_man = ref_man and self.ref_manufacturers[ref_man] or ''
            obj.edit(title = row['id'],
                     description = row.get('description', ''),
                     Blank = row['Blank'],
                     Hazardous = row['Hazardous'],
                     ReferenceResults = self.refsample_results[row['id']],
                     CatalogueNumber = row['CatalogueNumber'],
                     LotNumber = row['LotNumber'],
                     Remarks = row['Remarks'],
                     ExpiryDate = row['ExpiryDate'],
                     DateSampled = row['DateSampled'],
                     DateReceived = row['DateReceived'],
                     DateOpened = row['DateOpened'],
                     DateExpired = row['DateExpired'],
                     DateDisposed = row['DateDisposed']
                     )
            obj.setReferenceDefinition(ref_def)
            obj.setReferenceManufacturer(ref_man)
            obj.setCreators(row['creator'])
            obj.setCreationDate(row['created'])
            self.set_wf_history(obj, row['workflow_history'])
            self.ref_samples[row['id']] = obj
            obj.unmarkCreationFlag()
            validate(obj)

    def load_reference_analyses_interims(self, sheet):
        logger.info("Loading Reference Analyses Interim fields...")
        # Read the Ref Analyses interims self.refinterim_results
        self.refinterim_results = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            refanalysis_id = row['ReferenceAnalysis_id']
            if not self.refanalysis_results.has_key(refanalysis_id):
                self.refanalysis_results[refanalysis_id] = []
            self.refanalysis_results[refanalysis_id].append({
                'keyword': row['keyword'],
                'title': row['title'],
                'value': row['value']})

    def load_reference_analyses(self, sheet):
        logger.info("Loading Reference Analyses...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            service = self.services[row['AnalysisService_title']]
            # Analyses are keyed/named by service keyword
            sample = self.ref_samples[row['ReferenceSample_id']]
            sample.invokeFactory('ReferenceAnalysis', id = row['id'])
            obj = sample[row['id']]
            obj.edit(title = row['id'],
                     ReferenceType = row['ReferenceType'],
                     Result = row['Result'],
                     ResultDM = row['ResultDM'],
                     Analyst = row['Analyst'],
                     Instrument = row['Instrument'],
                     Retested = row['Retested']
                     )
            obj.setService(service)
            obj.setCreators(row['creator'])
            obj.setCreationDate(row['created'])
            self.set_wf_history(obj, row['workflow_history'])
            obj.unmarkCreationFlag()
            validate(obj)

    def load_attachment_types(self, sheet):
        logger.info("Loading Attachment Types...")
        folder = self.context.bika_setup.bika_attachmenttypes
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('AttachmentType', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_lab_products(self, sheet):
        logger.info("Loading Lab Products...")
        folder = self.context.bika_setup.bika_labproducts
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('LabProduct', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''),
                         Volume = row['Volume'],
                         Unit = row.get('Unit',''),
                         Price = "%02f" % float(row['Price']))
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_wst_layouts(self, sheet):
        logger.info("Loading Worksheet Layouts...")
        self.wst_layouts = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['WorksheetTemplate_title'] \
               not in self.wst_layouts.keys():
                self.wst_layouts[
                    row['WorksheetTemplate_title']] = []
            self.wst_layouts[
                row['WorksheetTemplate_title']].append({
                    'pos': row['pos'],
                    'type': row['type'][0].lower(),
                    'blank_ref': row['blank_ref'],
                    'control_ref': row['control_ref'],
                    'dup': row['dup']})

    def load_wst_services(self, sheet):
        logger.info("Loading Worksheet Services...")
        self.wst_services = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['WorksheetTemplate_title'] \
               not in self.wst_services.keys():
                self.wst_services[
                    row['WorksheetTemplate_title']] = []
            self.wst_services[
                row['WorksheetTemplate_title']].append(
                    self.services[row['service']].UID())

    def load_worksheet_templates(self, sheet):
        logger.info("Loading Worksheet Templates...")
        folder = self.context.bika_setup.bika_worksheettemplates
        self.wstemplates = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('WorksheetTemplate', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = row['title'],
                         description = row.get('description', ''),
                         Layout = self.wst_layouts[row['title']])
                obj.setService(self.wst_services[row['title']])
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)
                self.wstemplates[row['title']] = obj.UID()

    def load_partition_setup(self, sheet):
        logger.info("Loading Partition Setup...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['Analysis Service']: continue
            service = self.services[row['Analysis Service']]
            sampletype = self.sampletypes[row['Sample Type']]
            ps = service.getPartitionSetup()
            containers = []
            if row['Container']:
                for c in row['Container'].split(","):
                    c = c.strip()
                    containers.append(self.containers[c].UID())
            preservations = []
            if row['Preservation']:
                for p in row['Preservation'].split(","):
                    p = p.strip()
                    preservations.append(self.preservations[p].UID())
            ps.append({'sampletype': sampletype.UID(),
                       'container':containers,
                       'preservation':preservations,
                       'separate':row['Separate']})
            service.setPartitionSetup(ps)

    def load_samples(self, sheet):
        logger.info("Loading Samples...")
        self.samples = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            folder = self.clients[row['Client_Name']]
            _id = folder.invokeFactory('Sample', id = row['id'])
            obj = folder[_id]
            self.samples[row['id']] = obj.UID()
            obj.setSampleID(row['id'])
            obj.setClientSampleID(row['ClientSampleID'])
            obj.setSamplingWorkflowEnabled(False)
            obj.setDateSampled(row['DateSampled'])
            obj.setSampler(row['Sampler'])
            obj.setSamplingDate(row['SamplingDate'])
            obj.setDateReceived(row['DateReceived'])
            obj.setRemarks(row['Remarks'])
            obj.setComposite(row['Composite'])
            obj.setDateExpired(row['DateExpired'])
            obj.setDateDisposed(row['DateDisposed'])
            obj.setAdHoc(row['AdHoc'])
            obj.setSampleType(row['SampleType_title'])
            obj.setSamplePoint(row['SamplePoint_title'])
            if row['SamplingDeviation_title']:
                obj.setSamplingDeviation(self.samplingdeviations[row['SamplingDeviation_title']])
            obj.setCreators(row['creator'])
            obj.setCreationDate(row['created'])
            self.set_wf_history(obj, row['workflow_history'])
            review_state = self.wf.getInfoFor(obj, 'review_state')
            obj.unmarkCreationFlag()
            validate(obj)
            # Create a single partition...
            _id = obj.invokeFactory('SamplePartition', 'part-1')
            part = obj[_id]
            part.setContainer(self.containers['None Specified'])
            part.unmarkCreationFlag()
            validate(part)
            changeWorkflowState(part, 'bika_sample_workflow', review_state)
            part.reindexObject()

    def load_arccs(self, sheet):
        logger.info("Loading Analysis Request CCs")
        self.arccs = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['AnalysisRequest_id'] \
               not in self.arccs.keys():
                self.arccs[row['AnalysisRequest_id']] = []
            self.arccs[row['AnalysisRequest_id']].append(row['Contact_Fullname'])

    def load_ars(self, sheet):
        logger.info("Loading Analysis Requests")
        self.ars = {}
        # We apply the workflows -after- we add the analyses
        self.ar_workflows = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            folder = self.clients[row['Client_title']]
            _id = folder.invokeFactory('AnalysisRequest', id = row['id'])
            obj = folder[_id]
            self.ars[row['id']] = obj
            obj.edit(
                RequestID = row['id'],
                CCEmails = row['CCEmails'],
                ClientOrderNumber = row['ClientOrderNumber'],
                #Invoice
                InvoiceExclude = row['InvoiceExclude'],
                ReportDryMatter = row['ReportDryMatter'],
                DateReceived = row['DateReceived'],
                DatePublished = row['DatePublished'],
                Remarks = row['Remarks']
            )
            if row['Profile_title']:
                obj.setProfile(self.profiles[row['Profile_title']])
            if row['Template_title']:
                obj.setTemplate(self.artemplates[row['Template_title']])
            obj.setContact(self.client_contacts[row['Contact_Fullname']])
            obj.setSample(self.samples[row['Sample_id']])
            if self.arccs.get(row['id']):
                cc_contacts = [self.client_contacts[cc]
                               for cc in self.arccs.get(row['id'])]
                obj.setCCContact(cc_contacts)
            obj.setCreators(row['creator'])
            obj.setCreationDate(row['created'])
            self.ar_workflows[row['id']] = row['workflow_history']
            obj.unmarkCreationFlag()
            validate(obj)

    def load_analyses_interims(self, sheet):
        logger.info("Loading Analyses Interims...")
        # Read the AR Template Services into self.artemplate_analyses
        self.interim_results = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            analysis_id = row['Analysis_id']
            if not self.interim_results.has_key(analysis_id):
                self.interim_results[analysis_id] = []
            self.interim_results[analysis_id].append({
                'keyword': row['keyword'],
                'title': row['title'],
                'value': row['value']})

    def load_analyses(self, sheet):
        logger.info("Loading Analyses...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            service = self.services[row['Service_title']]
            # analyses are keyed/named by keyword
            keyword = service.getKeyword()
            ar = self.ars[row['AnalysisRequest_id']]
            ar.invokeFactory('Analysis', id = keyword)
            obj = ar[keyword]
            obj.edit(
                # Calculation = self.calculations # COU has none set
                # Attachment # COU has none
                Result = row['Result'],
                ResultCaptureDate = '', # COU has none.
                ResultDM = '', # COU has none.
                Retested = row['Retested'],
                MaxTimeAllowed = eval(row['MaxTimeAllowed']),
                DateAnalysisPublished = '', # XXX
                DueDate = row['DueDate'],
                Duration = row['Duration'],
                Earliness = row['Earliness'],
                ReportDryMatter = row['ReportDryMatter'],
                Analyst = row['Analyst'],
                Instrument = row['Instrument'],
                )
            # results are keyed by ID
            if row['id'] in self.interim_results.keys():
                InterimFields = self.interim_results[row['id']],
            part = obj.aq_parent.getSample().objectValues('SamplePartition')[0].UID()
            obj.setSamplePartition(part)
            obj.setService(service.UID())
            obj.setCreators(row['creator'])
            obj.setCreationDate(row['created'])
            self.set_wf_history(obj, row['workflow_history'])
            analyses = ar.objectValues('Analyses')
            analyses = list(analyses)
            analyses.append(obj)
            ar.setAnalyses(analyses)
            obj.unmarkCreationFlag()
            validate(obj)

    def load_worksheet_layouts(self, sheet):
        logger.info("Loading Worksheet Layouts...")
        self.ws_layouts = {}
        self.ws_analyses = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['Worksheet_id'] not in self.ws_layouts.keys():
                self.ws_layouts[row['Worksheet_id']] = []
                self.ws_analyses[row['Worksheet_id']] = []

            container = self.bc(id=row['container'])[0].getObject()
            container_uid = container.UID()
            try:
                analysis = container[row['analysis']]
            except:
                print "Can not find analysis %s in container %s: SKIP" % (row['analysis'], row['container'])
                continue
            analysis_uid = analysis.UID()

            self.ws_analyses[row['Worksheet_id']].append(analysis_uid)
            self.ws_layouts[row['Worksheet_id']].append({
                'position': row['position'],
                'type': row['type'],
                'container_uid': container_uid,
                'analysis_uid': analysis_uid,
            })

        for ws_id, wsl in self.ws_layouts.items():
            ws_uid = self.worksheets[row['Worksheet_id']]
            ws = self.uc(UID=ws_uid)[0].getObject()
            ws.setLayout(wsl)
            wsa = self.ws_analyses[ws_id]
            ws.setAnalyses(wsa)

    def load_worksheets(self, sheet):
        logger.info("Loading Worksheets...")
        self.worksheets = {}
        folder = self.context.worksheets
        rows = self.get_rows(sheet, 3)
        for row in rows:
            folder.invokeFactory('Worksheet', id = row['id'])
            obj = folder[row['id']]
            obj.edit(
                Remarks = row['Remarks'],
                Analyst = row['Analyst']
            )
            if row['WorksheetTemplate_title']:
                obj.setWorksheetTemplate(self.wstemplates[row['WorksheetTemplate_title']])
            obj.setCreators(row['creator'])
            obj.setCreationDate(row['created'])
            self.set_wf_history(obj, row['workflow_history'])
            self.worksheets[row['id']] = obj.UID()
            obj.unmarkCreationFlag()
            validate(obj)

    def load_duplicate_interims(self, sheet):
        logger.info("Loading Duplicate Interims...")
        self.d_interim_results = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            analysis_id = row['DuplicateAnalysis_id']
            if not self.d_interim_results.has_key(analysis_id):
                self.d_interim_results[analysis_id] = []
            self.d_interim_results[analysis_id].append({
                'keyword': row['keyword'],
                'title': row['title'],
                'value': row['value']})

    def load_duplicate_analyses(self, sheet):
        logger.info("Loading Duplicate Analyses...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            ws = self.uc(UID=self.worksheets[row['Worksheet_id']])[0].getObject()
            ws.invokeFactory('DuplicateAnalysis', id = row['id'])
            obj = ws[row['id']]
            obj.edit(
                InterimFields = self.d_interim_results.get(row['id'], []),
                Result = row['Result'],
                ResultDM = '', # XXX
                Retested = row['Retested'],
                Analyst = '', # XXX
                Instrument = '', # XXX
            )
            obj.setCreators(row['creator'])
            obj.setCreationDate(row['created'])
            self.set_wf_history(obj, row['workflow_history'])
            obj.unmarkCreationFlag()
            validate(obj)

    def load_manufacturers(self, sheet):
        logger.info("Loading Manufacturers...")
        folder = self.context.bika_setup.bika_manufacturers
        self.manufacturers = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            _id = folder.invokeFactory('Manufacturer', id = 'tmp')
            obj = folder[_id]
            if row['title']:
                obj.edit(title = row['title'],
                         description = row.get('description', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                self.manufacturers[row['title']] = obj.UID()
                renameAfterCreation(obj)

    def load_suppliers(self, sheet):
        logger.info("Loading Suppliers...")
        folder = self.context.bika_setup.bika_suppliers
        self.suppliers = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            _id = folder.invokeFactory('Supplier', id = 'tmp')
            obj = folder[_id]
            if row['Name']:
                obj.edit(AccountNumber = row.get('AccountNumber', ''),
                         Name = row['Name'],
                         EmailAddress = row['EmailAddress'],
                         Phone = row.get('Phone',''),
                         Fax = row.get('Fax',''))
                obj.unmarkCreationFlag()
                validate(obj)
                self.suppliers[row['Name']] = obj.UID()
                renameAfterCreation(obj)

    def load_supplier_contacts(self, sheet):
        logger.info("Loading Supplier Contacts...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['Supplier_Name']:
                continue
            if not row['Firstname']:
                continue
            folder = self.bsc(portal_type="Supplier",
                              Title=row['Supplier_Name'])
            if (len(folder) > 0):
                folder = folder[0].getObject()
                _id = folder.invokeFactory('SupplierContact', id='tmp')
                obj = folder[_id]
                obj.edit(
                    Firstname=row['Firstname'],
                    Surname=row.get('Surname', ''),
                    EmailAddress=row.get('EmailAddress', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

                if 'Username' in row:
                    obj.setUsername(row['Username'])

    def load_instrumenttypes(self, sheet):
        logger.info("Loading Instrument Types...")
        folder = self.context.bika_setup.bika_instrumenttypes
        self.instrumenttypes = {}
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if row['title']:
                _id = folder.invokeFactory('InstrumentType', id='tmp')
                obj = folder[_id]
                obj.edit(title=row['title'],
                         description=row.get('description', ''))
                obj.unmarkCreationFlag()
                validate(obj)
                self.instrumenttypes[row['title']] = obj.UID()
                renameAfterCreation(obj)

    def load_instrumentvalidations(self, sheet):
        logger.info("Loading Instrument Validations...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['instrument'] or not row['title']:
                continue

            folder = self.bsc(UID=self.instruments[row['instrument']])
            if len(folder) > 0:
                folder = folder[0].getObject()
                _id = folder.invokeFactory('InstrumentValidation', id='tmp')
                obj = folder[_id]
                obj.edit(
                         title=row['title'],
                         DownFrom=row.get('downfrom', ''),
                         DownTo=row.get('downto', ''),
                         Validator=row.get('validator', ''),
                         Considerations=row.get('considerations', ''),
                         WorkPerformed=row.get('workperformed', ''),
                         Remarks=row.get('remarks', '')
                         )
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_instrumentcalibrations(self, sheet):
        logger.info("Loading Instrument Calibrations...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['instrument'] or not row['title']:
                continue

            folder = self.bsc(UID=self.instruments[row['instrument']])
            if len(folder) > 0:
                folder = folder[0].getObject()
                _id = folder.invokeFactory('InstrumentCalibration', id='tmp')
                obj = folder[_id]
                obj.edit(
                         title=row['title'],
                         DownFrom=row.get('downfrom', ''),
                         DownTo=row.get('downto', ''),
                         Calibrator=row.get('calibrator', ''),
                         Considerations=row.get('considerations', ''),
                         WorkPerformed=row.get('workperformed', ''),
                         Remarks=row.get('remarks', '')
                         )
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_instrumentcertifications(self, sheet):
        logger.info("Loading Instrument Certifications...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['instrument'] or not row['title']:
                continue

            folder = self.bsc(UID=self.instruments[row['instrument']])
            if len(folder) > 0:
                folder = folder[0].getObject()
                _id = folder.invokeFactory('InstrumentCertification', id='tmp')
                obj = folder[_id]
                obj.edit(
                         title=row['title'],
                         Date=row.get('date', ''),
                         ValidFrom=row.get('validfrom', ''),
                         ValidTo=row.get('validto', ''),
                         Agency=row.get('agency', ''),
                         Remarks=row.get('remarks', '')
                         )
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_instrumentmaintenancetasks(self, sheet):
        logger.info("Loading Instrument Maintenance Tasks...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['instrument'] or not row['title'] or not row['type']:
                continue

            folder = self.bsc(UID=self.instruments[row['instrument']])
            if len(folder) > 0:
                folder = folder[0].getObject()
                _id = folder.invokeFactory('InstrumentMaintenanceTask',
                                           id='tmp')
                obj = folder[_id]
                obj.edit(
                         title=row['title'],
                         description=row['description'],
                         Type=row['type'],
                         DownFrom=row.get('downfrom', ''),
                         DownTo=row.get('downto', ''),
                         Maintainer=row.get('maintaner', ''),
                         Considerations=row.get('considerations', ''),
                         WorkPerformed=row.get('workperformed', ''),
                         Remarks=row.get('remarks', ''),
                         Cost=row.get('cost', ''),
                         Closed=self.to_bool(row.get('closed'))
                         )
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def load_instrumentschedule(self, sheet):
        logger.info("Loading Instrument Schedule...")
        rows = self.get_rows(sheet, 3)
        for row in rows:
            if not row['instrument'] or not row['title'] or not row['type']:
                continue

            folder = self.bsc(UID=self.instruments[row['instrument']])
            if len(folder) > 0:
                folder = folder[0].getObject()
                _id = folder.invokeFactory('InstrumentScheduledTask',
                                           id='tmp')
                criteria = [{'fromenabled':row.get('date', None) is not None,
                             'fromdate':row.get('date', ''),
                             'repeatenabled':((row['numrepeats'] and
                                               row['numrepeats'] > 1) or
                                              (row['repeatuntil'] and
                                               len(row['repeatuntil']) > 0)),
                             'repeatunit':row.get('numrepeats', ''),
                             'repeatperiod':row.get('periodicity', ''),
                             'repeatuntilenabled':(row['repeatuntil'] and
                                                len(row['repeatuntil']) > 0),
                             'repeatuntil':row.get('repeatuntil')}]
                obj = folder[_id]
                obj.edit(
                         title=row['title'],
                         Type=row['type'],
                         ScheduleCriteria=criteria,
                         Considerations=row.get('considerations', ''),
                         )
                obj.unmarkCreationFlag()
                validate(obj)
                renameAfterCreation(obj)

    def get_rows(self, sheet, startrow=3):

        """ Returns an array with all the data from a sheet.
            Each row contains a dictionary where the key is the value of the
            first row of the sheet for each column.
            The data values are returned in utf-8 format.
            Starts to consume data from startrow
        """

        rowsout = []
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in
                 range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[0]
        for row in rows[startrow:]:
            rowout = [_c(r).decode('utf-8') for r in row]
            rowout = dict(zip(fields, rowout))
            rowsout.append(rowout)
        return rowsout

    def to_bool(self, value):

        """ Converts a sheet string value to a boolean value.
            Needed because of utf-8 conversions
        """

        if value is not None and (value == u'True' or value == u'true' \
            or value == 'True' or value == 'true' \
            or value == u'1' or value == '1' or value == 1):
            return True
        else:
            return False
