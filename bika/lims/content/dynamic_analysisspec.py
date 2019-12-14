# -*- coding: utf-8 -*-

from collections import defaultdict
from StringIO import StringIO

from bika.lims import _
from bika.lims.catalog import SETUP_CATALOG
from openpyxl.reader.excel import load_workbook
from openpyxl.shared.exc import InvalidFileException
from plone.dexterity.content import Item
from plone.namedfile import field as namedfile
from plone.supermodel import model
from zope.interface import Invalid
from zope.interface import implementer
from zope.interface import invariant

REQUIRED_COLUMNS = [
    "Keyword",  # The Analysis Keyword
    "Min",  # Lower Limit
    "Max",  # Upper Limit
]


class IDynamicAnalysisSpec(model.Schema):
    """Dynamic Analysis Specification
    """

    specs_file = namedfile.NamedBlobFile(
        title=_(u"Specification File"),
        description=_(u"Only Excel files supported"),
        required=True)

    @invariant
    def validate_sepecs_file(data):
        """Checks the Excel file contains the required header columns
        """
        fd = StringIO(data.specs_file.data)
        try:
            xls = load_workbook(fd)
        except (InvalidFileException, TypeError):
            raise Invalid(_(
                "Invalid specifications file detected. "
                "Please upload an Excel spreadsheet with at least "
                "the following columns defined: '{}'"
                .format(", ".join(REQUIRED_COLUMNS))))
        try:
            header = map(lambda c: c.value, xls.worksheets[0].rows[0])
        except IndexError:
            raise Invalid(
                _("First sheet does not contain a valid column definition"))
        for col in REQUIRED_COLUMNS:
            if col not in header:
                raise Invalid(_("Column '{}' is missing".format(col)))


@implementer(IDynamicAnalysisSpec)
class DynamicAnalysisSpec(Item):
    """Dynamic Analysis Specification
    """
    _catalogs = [SETUP_CATALOG]

    def get_workbook(self):
        specs_file = self.specs_file
        if not specs_file:
            return None
        data = StringIO(specs_file.data)
        return load_workbook(data)

    def get_worksheets(self):
        wb = self.get_workbook()
        if wb is None:
            return []
        return wb.worksheets

    def get_primary_sheet(self):
        sheets = self.get_worksheets()
        if len(sheets) == 0:
            return None
        return sheets[0]

    def get_header(self):
        ps = self.get_primary_sheet()
        if ps is None:
            return []
        return map(lambda cell: cell.value, ps.rows[0])

    def get_specs(self):
        ps = self.get_primary_sheet()
        if ps is None:
            return []
        keys = self.get_header()
        specs = []
        for num, row in enumerate(ps.rows):
            # skip the header
            if num == 0:
                continue
            values = map(lambda cell: cell.value, row)
            data = dict(zip(keys, values))
            specs.append(data)
        return specs

    def get_by_keyword(self):
        specs = self.get_specs()
        groups = defaultdict(list)
        for spec in specs:
            groups[spec.get("Keyword")].append(spec)
        return groups
