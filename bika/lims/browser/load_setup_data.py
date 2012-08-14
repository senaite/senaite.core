from App.Common import package_home
from Products.Archetypes.config import REFERENCE_CATALOG
from Products.CMFCore.utils import getToolByName
from Products.Five.browser import BrowserView
from Products.Five.browser.pagetemplatefile import ViewPageTemplateFile
from bika.lims import bikaMessageFactory as _
from bika.lims import PMF
from bika.lims import logger
from bika.lims.idserver import renameAfterCreation
from bika.lims.utils import sortable_title
from cStringIO import StringIO
from openpyxl.reader.excel import load_workbook
from os.path import join
from xml.etree.ElementTree import XML
from zipfile import ZipFile, ZIP_DEFLATED
from zope.app.component.hooks import getSite
import Globals
import tempfile
import transaction
import zope
from pkg_resources import resource_listdir, resource_filename, ResourceManager

class LoadSetupData(BrowserView):
    template = ViewPageTemplateFile("templates/load_setup_data.pt")

    def __init__(self, context, request):
        BrowserView.__init__(self, context, request)
        self.title = _("Load Setup Data")
        self.description = _("Submit a valid Open XML (.XLSX) file containing Bika setup records to continue.")
        # dependencies to resolve
        self.deferred = {}

        self.request.set('disable_border', 1)

    def getSetupDatas(self):
        files = []
        for f in resource_listdir('bika.lims', 'setupdata'):
            f = f.lower()
            if f.find(".xlsx") > -1 and f.find("_testing") == -1:
                files.append(f.replace(".xlsx",""))
        return files

    def __call__(self):
        form = self.request.form

        if not 'submitted' in form:
            return self.template()

        self.portal_catalog = getToolByName(self.context, 'portal_catalog')
        self.bc = bsc = getToolByName(self.context, 'bika_catalog')
        self.bsc = bsc = getToolByName(self.context, 'bika_setup_catalog')
        self.reference_catalog = getToolByName(self.context, REFERENCE_CATALOG)
        self.portal_registration = getToolByName(self.context, 'portal_registration')
        self.portal_groups = getToolByName(self.context, 'portal_groups')
        self.portal_membership = getToolByName(self.context, 'portal_membership')
        self.plone_utils = getToolByName(self.context, 'plone_utils')
        portal = getSite()

        file_name = 'xlsx' in form and form['xlsx'] or None
        if not file_name:
            msg = self.context.translate(
                _("No file data submitted.  Please submit a valid Open "
                  "XML Spreadsheet (.xlsx) file."))
            self.plone_utils.addPortalMessage(msg)
            return self.template()

        self.file_name = file_name

        filename = resource_filename("bika.lims",
                                     "setupdata/%s.xlsx"%file_name)
        wb = load_workbook(filename = filename)

        sheets = {}
        for sheetname in wb.get_sheet_names():
            sheets[sheetname] = wb.get_sheet_by_name(sheetname)
        self.load_lab_information(sheets['Lab Information'])
        self.load_lab_users(sheets['Lab Users'])
        self.load_lab_contacts(sheets['Lab Contacts'])
        self.load_lab_departments(sheets['Lab Departments'])
        self.load_containertypes(sheets["Container Types"])
        self.load_preservations(sheets["Preservations"])
        self.load_containers(sheets["Containers"])
        self.load_clients(sheets['Clients'])
        self.client_contacts = []
        self.load_client_contacts(sheets['Client Contacts'])
        self.fix_client_contact_ccs()
        self.instruments = {}
        self.load_instruments(sheets['Instruments'])
        self.load_sample_matrices(sheets['Sample Matrices'])
        self.load_sample_points(sheets['Sample Points'])
        self.load_sample_types(sheets['Sample Types'])
        self.cats = {}
        self.load_analysis_categories(sheets['Analysis Categories'])
        self.load_methods(sheets['Methods'])
        self.calcs = {}
        self.services = {}
        self.load_analysis_services(sheets['Analysis Services'])
        self.load_calculations(sheets['Calculations'])
        self.load_partition_setup(sheets['Partition Setup'])

        # process deferred services and calculations which depend on each other
        nr_deferred = 0
        while self.deferred['Analysis Services'] or \
              self.deferred['Calculations']:
            current_deferred = len(self.deferred['Calculations']) + \
                               len(self.deferred['Analysis Services'])
            if (self.deferred['Calculations'] or \
                self.deferred['Analysis Services']) and \
               nr_deferred == current_deferred:
                msg = "The following dependencies are unsatisfied:<br>\n"
                for service in self.deferred['Analysis Services']:
                    msg += "Service: %s -> Calculation: %s<br>\n" % \
                        (service['title'], service['Calculation'])
                for calculation in self.deferred['Calculations']:
                    msg += "Calculation: %s -> Services: %s<br>\n" % \
                        (calculation['title'], calculation['DependentServices'])
                raise Exception,msg
            nr_deferred = current_deferred
            if self.deferred['Analysis Services']:
                defs = self.deferred['Analysis Services']
                self.deferred['Analysis Services'] = []
                self.CreateServiceObjects(defs)
            if self.deferred['Calculations']:
                defs = self.deferred['Calculations']
                self.deferred['Calculations'] = []
                self.CreateCalculationObjects(defs)
            if not self.deferred['Calculations'] and \
               not self.deferred['Analysis Services']:
                break

        self.load_analysis_specifications(sheets['Analysis Specifications'])
        self.load_analysis_profiles(sheets['Analysis Profiles'])
        self.load_reference_definitions(sheets['Reference Definitions'])
        self.load_reference_suppliers(sheets['Reference Suppliers'])
        self.load_reference_supplier_contacts(sheets['Reference Supplier Contacts'])
        self.load_attachment_types(sheets['Attachment Types'])
        self.load_lab_products(sheets['Lab Products'])
        self.load_sampling_deviations(sheets['Sampling Deviations'])
        self.load_worksheet_templates(sheets['Worksheet Templates'])
        self.load_reference_manufacturers(sheets['Reference Manufacturers'])
        self.load_bika_setup(sheets['Setup'])

        bsc = getToolByName(self.context, 'bika_setup_catalog')
        bsc.clearFindAndRebuild()

        message = PMF("Changes saved.")
        self.plone_utils.addPortalMessage(message)
        self.request.RESPONSE.redirect(portal.absolute_url())

    def load_bika_setup(self,sheet):
        nr_rows = sheet.get_highest_row()
##        self.request.response.write("<input type='hidden' id='load_section' value='BikaSetup' max='%s'/>"%(nr_rows))
##        self.request.response.flush()
        values = {}
        for r in range(nr_rows):
            values[sheet.cell(row=r, column=0).value] = \
                sheet.cell(row=r, column=2).value
        self.context.bika_setup.edit(
            PasswordLifetime = int(values['PasswordLifetime']),
            AutoLogOff = int(values['AutoLogOff']),
            Currency = values['Currency'],
            MemberDiscount = str(float(values['MemberDiscount'])),
            VAT = str(float(values['VAT'])),
            MinimumResults = int(values['MinimumResults']),
            BatchEmail = int(values['BatchEmail']),
            BatchFax = int(values['BatchFax']),
            SMSGatewayAddress = values['SMSGatewayAddress'],
            SamplingWorkflowEnabled = values['SamplingWorkflowEnabled'],
            CategoriseAnalysisServices = values['CategoriseAnalysisServices'],
            DryMatterService = self.services[values['DryMatterService']],
            ARImportOption = values['ARImportOption'],
            ARAttachmentOption = values['ARAttachmentOption'],
            AnalysisAttachmentOption = values['AnalysisAttachmentOption'],
            DefaultSampleLifetime = eval(values['DefaultSampleLifetime']),
            AutoPrintLabels = values['AutoPrintLabels'],
            AutoLabelSize = values['AutoLabelSize'],
            YearInPrefix = values['YearInPrefix'],
            SampleIDPadding = int(values['SampleIDPadding']),
            ARIDPadding = int(values['ARIDPadding']),
            ExternalIDServer = values['ExternalIDServer'],
            IDServerURL = values['IDServerURL'],
        )

    def load_containertypes(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='ContainerTypes' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_containertypes
        self.containertypes = {}
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('ContainerType', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)
            self.containertypes[unicode(row['title'])] = obj


    def load_preservations(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Preservations' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_preservations
        self.preservations = {}
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('Preservation', id = 'tmp')
            obj = folder[_id]
            containertypes = []
            if row['ContainerType']:
                for ct in row['ContainerType'].split(","):
                    containertypes.append(self.containertypes[ct.strip()])
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     RetentionPeriod = row['RetentionPeriod'] and eval(row['RetentionPeriod']) or {},
                     ContainerType = containertypes
                     )
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)
            self.preservations[unicode(row['title'])] = obj


    def load_containers(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Containers' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_containers
        self.containers = {}
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('Container', id = 'tmp')
            obj = folder[_id]
            P = row['Preservation']
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     Capacity = unicode(row['Capacity']),
                     ContainerType = row['ContainerType'] \
                         and self.containertypes[row['ContainerType']] or None,
                     PrePreserved = row['PrePreserved'] and row['PrePreserved'] or False,
                     Preservation = P and self.preservations[P] or None)
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)
            self.containers[unicode(row['title'])] = obj


    def load_lab_information(self, sheet):
        self.departments = {}
        laboratory = self.context.bika_setup.laboratory
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Lab Information' max='1'/>")
##        self.request.response.flush()

        rows = [[sheet.cell(row=row_nr, column=col_nr).value
                for col_nr in range(nr_cols + 1)]
                        for row_nr in range(nr_rows)]

        lab_info = {}
        for row in rows:
            lab_info[row[2]] = row[1]

        laboratory.edit(
            Name = unicode(lab_info['Name']),
            TaxNumber = unicode(lab_info['TaxNumber']),
            Phone = unicode(lab_info['Phone']),
            EmailAddress = unicode(lab_info['EmailAddress']),
            LabURL = unicode(lab_info['LabURL']),
            Confidence = unicode(lab_info['Confidence']),
            LaboratoryAccredited = lab_info['LaboratoryAccredited'],
            AccreditationBody = unicode(lab_info['AccreditationBody']),
            AccreditationBodyLong = unicode(lab_info['AccreditationBodyLong']),
            Accreditation = unicode(lab_info['Accreditation']),
            AccreditationReference = unicode(lab_info['AccreditationReference']),
            PhysicalAddress = unicode(lab_info['PhysicalAddress'])),

    def load_lab_users(self, sheet):
        portal_registration = getToolByName(self.context, 'portal_registration')
        portal_groups = getToolByName(self.context, 'portal_groups')
        portal_membership = getToolByName(self.context, 'portal_membership')
        plone_utils = getToolByName(self.context, 'plone_utils')

        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Lab Users' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value
                 for col_nr in range(nr_cols)]
                  for row_nr in range(nr_rows)]
        fields = rows[1]
        for row in rows[3:]:
            row = dict(zip(fields, row))
            portal_registration.addMember(
                unicode(row['Username']),
                unicode(row['Password']),
                properties = {
                    'username': unicode(row['Username']),
                    'email': unicode(row['EmailAddress']),
                    'fullname': " ".join((row['Firstname'], row['Surname']))})
            group_ids = [g.strip() for g in unicode(row['Groups']).split(',')]
            for group_id in group_ids:
                group = portal_groups.getGroupById(group_id)
                if not group:
                    message = self.context.translate(
                        "message_invalid_group",
                        "bika",
                        {'group_id': group_id},
                        self.context,
                        default = "Invalid group: '${group_id}'.")
                    plone_utils.addPortalMessage(message)
                    return self.template()
                group.addMember(unicode(row['Username']))
            # If user is in LabManagers, add Owner local role on clients folder
            if 'LabManager' in group_ids:
                self.context.clients.manage_setLocalRoles(unicode(row['Username']),
                                                          ['Owner',] )

    def load_lab_contacts(self, sheet):
        self.lab_contacts = []
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Lab Contacts' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_labcontacts
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('LabContact', id='tmp')
            obj = folder[_id]
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)
            Fullname = unicode(row['Firstname']) + " " + unicode(row['Surname'])
            obj.edit(
                title = Fullname,
                description = Fullname,
                Firstname = unicode(row['Firstname']),
                Surname = unicode(row['Surname']),
                EmailAddress = unicode(row['EmailAddress']),
                BusinessPhone = unicode(row['BusinessPhone']),
                BusinessFax = unicode(row['BusinessFax']),
                MobilePhone = unicode(row['MobilePhone']),
                JobTitle = unicode(row['JobTitle']))
                # Department = row['Department'],
                # Signature = row['Signature'],
            row['obj'] = obj
            self.lab_contacts.append(row)

    def load_lab_departments(self, sheet):
        self.departments = {}
        lab_contacts = self.bsc(portal_type="LabContact")
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Lab Departments' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_departments
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('Department', id = 'tmp')
            obj = folder[_id]
            manager = None
            for contact in lab_contacts:
                contact = contact.getObject()
                if contact.getFullname() == unicode(row['_LabContact_Fullname']):
                    manager = contact
                    break
            if not manager:
                message = "Error: lookup of '%s' in LabContacts/Fullname failed."%unicode(row['_LabContact_Fullname'])
                self.plone_utils.addPortalMessage(message)
                raise Exception(message)
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     Manager = manager.UID())
            self.departments[unicode(row['title'])] = obj
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

            # set importedlab contact's department references
            if hasattr(self, 'lab_contacts'):
                for contact in self.lab_contacts:
                    if contact['Department'] == unicode(row['title']):
                        contact['obj'].setDepartment(obj.UID())


    def load_clients(self, sheet):
        self.clients = []
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Clients' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.clients
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('Client', id = 'tmp')
            obj = folder[_id]
            obj.edit(AccountNumber = unicode(row['AccountNumber']),
                     Name = unicode(row['Name']),
                     MemberDiscountApplies = row['MemberDiscountApplies'] and True or False,
                     EmailAddress = unicode(row['EmailAddress']),
                     Phone = unicode(row['Telephone']),
                     Fax = unicode(row['Fax']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)


    def load_client_contacts(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Client Contacts' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.clients
        for row in rows[3:]:
            row = dict(zip(fields, row))
            client = self.portal_catalog(portal_type = "Client",
                                         Title = unicode(row['_Client_Name']))
            if len(client) == 0:
                raise IndexError("Client invalid: '%s'" % unicode(row['_Client_Name']))
            client = client[0].getObject()
            _id = client.invokeFactory('Contact', id = 'tmp')
            contact = client[_id]
            cc = self.portal_catalog(portal_type="Contact",
                                     getUsername = [c.strip() for c in unicode(row['CC']).split(',')])
            if row['CC'] and not cc:
                row['uid'] = contact.UID()
                self.client_contacts.append(row)
            contact.edit(Salutation = unicode(row['Salutation']),
                         Firstname = unicode(row['Firstname']),
                         Surname = unicode(row['Surname']),
                         JobTitle = unicode(row['JobTitle']),
                         Department = unicode(row['Department']),
                         BusinessPhone = unicode(row['BusinessPhone']),
                         BusinessFax = unicode(row['BusinessFax']),
                         HomePhone = unicode(row['HomePhone']),
                         MobilePhone = unicode(row['MobilePhone']),
                         EmailAddress = unicode(row['EmailAddress']),
                         PublicationPreference = unicode(row['PublicationPreference']),
                         CCContact = [c.UID for c in cc],
                         AttachmentsPermitted = row['AttachmentsPermitted'] and True or False)
            if 'Username' in row and \
               'EmailAddress' in row and \
               'Password' in row:
                self.context.REQUEST.set('username', unicode(row['Username']))
                self.context.REQUEST.set('password', unicode(row['Password']))
                self.context.REQUEST.set('email', unicode(row['EmailAddress']))
                pr = getToolByName(self.context, 'portal_registration')
                try:
                    pr.addMember(unicode(row['Username']),
                                 unicode(row['Password']),
                                 properties = {
                                     'username': unicode(row['Username']),
                                     'email': unicode(row['EmailAddress']),
                                     'fullname': " ".join((unicode(row['Firstname']),
                                                           unicode(row['Surname'])))})
                    contact.setUsername(unicode(row['Username']))
                except ValueError:
                    logger.info("username %s is already in use" % unicode(row['Username']))
                    raise
                # Give contact's user an Owner local role on their client
                contact.aq_parent.manage_setLocalRoles(unicode(row['Username']),
                                                       ["Owner",])
                # add user to Clients group
                group = self.context.portal_groups.getGroupById('Clients')
                group.addMember(row['Username'])

            contact.unmarkCreationFlag()
            renameAfterCreation(contact)

    def fix_client_contact_ccs(self):
        for row in self.client_contacts:
            client = self.portal_catalog(portal_type = "Client",
                                         Title = unicode(row['_Client_Name']))
            contact = self.reference_catalog.lookupObject(row['uid'])
            cc = self.portal_catalog(portal_type="Contact",
                                     getUsername = [c.strip() for c in
                                                    unicode(row['CC']).split(',')])
            if cc:
                contact.setCCContact([c.UID for c in cc])

    def load_instruments(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Instruments' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_instruments
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('Instrument', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     Type = unicode(row['Type']),
                     Brand = unicode(row['Brand']),
                     Model = unicode(row['Model']),
                     SerialNo = unicode(row['SerialNo']),
                     CalibrationCertificate = unicode(row['CalibrationCertificate']),
                     CalibrationExpiryDate = unicode(row['CalibrationExpiryDate']),
                     DataInterface = row['DataInterface'])
            self.instruments[unicode(row['title'])] = obj
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)


    def load_sample_points(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Sample Points' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        setup_folder = self.context.bika_setup.bika_samplepoints
        for row in rows[3:]:
            row = dict(zip(fields, row))

            if row['_Client_Name']:
                client_name = unicode(row['_Client_Name'])
                client = self.portal_catalog(portal_type = "Client",
                                             Title = client_name)
                if len(client) == 0:
                    raise IndexError("Client invalid: '%s'" % client_name)
                folder = client[0].getObject()
            else:
                folder = setup_folder

            _id = folder.invokeFactory('SamplePoint', id = 'tmp')
            obj = folder[_id]
            latitude = {'degrees': row['lat deg'],
                        'minutes': row['lat min'],
                        'seconds': row['lat sec'],
                        'bearing': row['NS']}
            longitude = {'degrees': row['long deg'],
                        'minutes': row['long min'],
                        'seconds': row['long sec'],
                        'bearing': row['EW']}
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     SamplingFrequency = {'days':unicode(row['Days']), 'hours':unicode(row['Hours']), 'minutes':unicode(row['Minutes'])},
                     Latitude = latitude,
                     Longitude = longitude,
                     Elevation = unicode(row['Elevation']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

    def load_sample_types(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Sample Types' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_sampletypes
        self.sampletypes = {}
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('SampleType', id = 'tmp')
            obj = folder[_id]
            self.sampletypes[row['title']] = obj
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     RetentionPeriod = row['RetentionPeriod'] and eval(row['RetentionPeriod']) or {},
                     Prefix = unicode(row['Prefix']),
                     MinimumVolume = unicode(row['MinimumVolume']),
                     Hazardouus = row['Hazardous'] and True or False)
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

    def load_sampling_deviations(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Sampling Deviations' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_samplingdeviations
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('SamplingDeviation', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

    def load_sample_matrices(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Sample Matrices' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_samplematrices
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('SampleMatrix', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

    def load_analysis_categories(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Analysis Categories' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_analysiscategories
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('AnalysisCategory', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     Department = row['Department'] and self.departments[unicode(row['Department'])].UID() or None)
            self.cats[unicode(row['title'])] = obj
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

    def load_methods(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Methods' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.methods
        self.methods = {}
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('Method', id = 'tmp')
            obj = folder[_id]

            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     Instructions = unicode(row['Instructions']))

            if row['MethodDocument']:
                file_title = sortable_title(obj, row['MethodDocument'])
                path = resource_filename("bika.lims",
                                         "setupdata/%s/methods/%s" \
                                         % (self.file_name, row['MethodDocument']))
                #file_id = obj.invokeFactory("File", id=row['MethodDocument'])
                #thisfile = obj[file_id]
                file_data = open(path, "rb").read()
                obj.setMethodDocument(file_data)

            obj.unmarkCreationFlag()
            renameAfterCreation(obj)
            self.methods[unicode(row['title'])] = obj

    def CreateServiceObjects(self, services):
        deferred = 0
        if not hasattr(self, 'service_objs'):
            self.service_objs = {}
        if not self.deferred.has_key('Analysis Services'):
            self.deferred['Analysis Services'] = []

        folder = self.context.bika_setup.bika_analysisservices

        for row in services:
            if row['Calculation'] and not row['Calculation'] in \
               [c.Title() for c in self.calcs.values()]:
                self.deferred['Analysis Services'].append(row)
                deferred = 1
                continue
            deferred = 0

            _id = folder.invokeFactory('AnalysisService', id = 'tmp')
            obj = folder[_id]
            obj.edit(
                title = unicode(row['title']),
                description = row['description'] and unicode(row['description']) or '',
                Method = row['Method'] and self.methods[unicode(row['Method'])] or None,
                Container = row['Container'] and [self.containers[c] for c in row['Container'].split(",")] or [],
                Preservation = row['Preservation'] and [self.preservations[c] for c in row['Preservation'].split(",")] or [],
                PointOfCapture = unicode(row['PointOfCapture']),
                Unit = row['Unit'] and unicode(row['Unit']) or None,
                Category = self.cats[unicode(row['Category'])].UID(),
                Price = "%02f" % float(row['Price']),
                CorporatePrice = "%02f" % float(row['BulkPrice']),
                VAT = "%02f" % float(row['VAT']),
                Precision = unicode(row['Precision']),
                Accredited = row['Accredited'] and True or False,
                Keyword = unicode(row['Keyword']),
                MaxTimeAllowed = row['MaxTimeAllowed'] and eval(row['MaxTimeAllowed']) or {},
                DuplicateVariation = "%02f" % float(row['DuplicateVariation']),
                Uncertanties = row['Uncertainties'] and eval(row['Uncertainties']) or [],
                ResultOptions = row['ResultOptions'] and eval(row['ResultOptions']) or [],
                ReportDryMatter = row['ReportDryMatter'] and True or False
            )
            if row['Instrument']:
                obj.setInstrument(row['Instrument'] in self.instruments and self.instruments[row['Instrument']].UID()),
            if row['Calculation']:
                obj.setCalculation(self.calcs[row['Calculation']])
            service_obj = obj
            self.services[row['Keyword']] = obj
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

    def load_analysis_services(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Analysis Categories' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_analysisservices
        services = []
        for row in rows[3:]:
            row = dict(zip(fields, row))
            services.append(row)
        self.CreateServiceObjects(services)

    def CreateCalculationObjects(self, calcs):
        deferred = 0
        if not self.deferred.has_key('Calculations'):
            self.deferred['Calculations'] = []
        folder = self.context.bika_setup.bika_calculations
        for row in calcs:
            if not row['title']:
                if deferred:
                    i = {'keyword': unicode(row['interim_keyword']),
                          'title': unicode(row['interim_title']),
                          'type': unicode(row['interim_type']),
                          'value': unicode(row['interim_value']),
                          'unit': unicode(row['interim_unit'] and row['interim_unit'] or '')}
                    self.deferred['Calculations'][-1]['_interim'].append(i)
                    continue
                if row['interim_keyword']:
                    i = [{'keyword': unicode(row['interim_keyword']),
                          'title': unicode(row['interim_title']),
                          'type': unicode(row['interim_type']),
                          'value': unicode(row['interim_value']),
                          'unit': unicode(row['interim_unit'] and row['interim_unit'] or '')}]
                    calc_obj.setInterimFields(
                        calc_obj.getInterimFields() + i
                    )
                continue
            deps = row['DependentServices'] and \
                [d.strip() for d in unicode(row['DependentServices']).split(",")] or []
            depservices = []
            try:
                depservices = [self.services[d].UID() for d in deps]
            except KeyError:
                pass
            if len(deps) != len(depservices):
                # remaining interims in subsequent rows for deferred
                # calculations go in row/_interim
                if not '_interim' in row:
                    row['_interim'] = []
                self.deferred['Calculations'].append(row)
                deferred = 1
                continue
            deferred = 0
            _id = folder.invokeFactory('Calculation', id = 'tmp')
            obj = folder[_id]
            if row['interim_keyword']:
                i = [{'keyword': unicode(row['interim_keyword']),
                      'title': unicode(row['interim_title']),
                      'type': unicode(row['interim_type']),
                      'value': unicode(row['interim_value']),
                      'unit': unicode(row['interim_unit'] and row['interim_unit'] or '')}]
            else:
                i = []
            obj.edit(title = unicode(row['title']),
                      description = unicode(row['description']),
                      DependentServices = depservices,
                      InterimFields = i,
                      Formula = str(row['Formula']))
            if '_interim' in row:
                obj.setInterimFields(obj.getInterimFields() + row['_interim'])
            calc_obj = obj
            self.calcs[row['title']] = obj
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)


    def load_calculations(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Calculations' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_analysisservices
        calcs = []
        for row in rows[3:]:
            row = dict(zip(fields, row))
            calcs.append(row)
        self.CreateCalculationObjects(calcs)

    def load_analysis_profiles(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Analysis Profiles' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_analysisprofiles
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('AnalysisProfile', id = 'tmp')
            obj = folder[_id]
            services = [d.strip() for d in unicode(row['Service']).split(",")]
            proxies = self.bsc(portal_type="AnalysisService",
                               getKeyword = services)
            if len(proxies) != len(services):
                raise Exception("Analysis Profile services invalid.  Got %s, found %s" %\
                                (services, [p.getKeyword for p in proxies]))

            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     Service = [s.UID for s in proxies],
                     ProfileKey = unicode(row['ProfileKey']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)


    def load_reference_definitions(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Reference Definitions' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        self.definitions = {}
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_referencedefinitions
        for row in rows[3:]:
            row = dict(zip(fields, row))
            if row['title']:
                _id = folder.invokeFactory('ReferenceDefinition', id = 'tmp')
                obj = folder[_id]
                obj.edit(title = unicode(row['title']),
                         description = unicode(row['description']),
                         Blank = row['Blank'] and True or False,
                         Hazardous = row['Hazardous'] and True or False)
                obj.unmarkCreationFlag()
                renameAfterCreation(obj)
                self.definitions[unicode(row['title'])] = obj.UID()
            service = self.services[row['keyword']]
            try: result = int(row['result'])
            except: result = 0
            try: Min = int(row['min'])
            except: Min = 0
            try: Max = int(row['max'])
            except: Max = 0
            obj.setReferenceResults(
                obj.getReferenceResults() + [{'uid': service.UID(),
                                              'result':unicode(result),
                                              'min':unicode(Min),
                                              'max':unicode(Max)}])

    def load_analysis_specifications(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Analysis Specifications' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_analysisspecs
        ResultsRange = []
        for row in rows[3:]:
            row = dict(zip(fields, row))
            if row['SampleType']:
                if ResultsRange:
                    obj.setResultsRange(ResultsRange)
                    ResultsRange = []
                _id = folder.invokeFactory('AnalysisSpec', id = 'tmp')
                obj = folder[_id]
                SampleType = self.sampletypes[row['SampleType']]
                obj.edit(SampleType = SampleType.UID(),
                         title = row['SampleType'])
                obj.unmarkCreationFlag()
                renameAfterCreation(obj)
            else:
                ResultsRange.append({'keyword': row['keyword'],
                                     'min': str(row['min']),
                                     'max': str(row['max']),
                                     'error': str(row['error'])},)

    def load_reference_suppliers(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Reference Suppliers' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_referencesuppliers
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('ReferenceSupplier', id = 'tmp')
            obj = folder[_id]
            obj.edit(AccountNumber = unicode(row['AccountNumber']),
                     Name = unicode(row['Name']),
                     EmailAddress = unicode(row['EmailAddress']),
                     Phone = unicode(row['Phone']),
                     Fax = unicode(row['Fax']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)


    def load_reference_supplier_contacts(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Reference Supplier Contacts' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        for row in rows[3:]:
            row = dict(zip(fields, row))
            if not row['_ReferenceSupplier_Name']:
                continue
            folder = self.bsc(portal_type="ReferenceSupplier",
                              Title = row['_ReferenceSupplier_Name'])[0].getObject()
            _id = folder.invokeFactory('SupplierContact', id = 'tmp')
            obj = folder[_id]
            obj.edit(Firstname = unicode(row['Firstname']),
                     Surname = unicode(row['Surname']),
                     EmailAddress = unicode(row['EmailAddress']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

            if 'Username' in row and \
               'Password' in row:
                self.context.REQUEST.set('username', unicode(row['Username']))
                self.context.REQUEST.set('password', unicode(row['Password']))
                self.context.REQUEST.set('email', unicode(row['EmailAddress']))
                pr = getToolByName(self.context, 'portal_registration')
                pr.addMember(unicode(row['Username']),
                             unicode(row['Password']),
                             properties = {
                                 'username': unicode(row['Username']),
                                 'email': unicode(row['EmailAddress']),
                                 'fullname': " ".join((row['Firstname'],
                                                       row['Surname']))})
                obj.setUsername(unicode(row['Username']))

    def load_attachment_types(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Attachment Types' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_attachmenttypes
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('AttachmentType', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)


    def load_lab_products(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Lab Products' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_labproducts
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('LabProduct', id = 'tmp')
            obj = folder[_id]

            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     Volume = unicode(row['Volume']),
                     Unit = unicode(row['Unit'] and row['Unit'] or ''),
                     Price = "%02f" % float(row['Price']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

    def load_worksheet_templates(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Worksheet Templates' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_worksheettemplates
        for row in rows[3:]:
            row = dict(zip(fields, row))
            if not row['title']:
                if row['pos']:
                    control_ref = self.definitions.get(unicode(row['control_ref']), '')
                    blank_ref = self.definitions.get(unicode(row['blank_ref']), '')
                    l = [{'pos':unicode(row['pos']),
                          'type':unicode(row['type']),
                          'control_ref':control_ref,
                          'blank_ref':blank_ref,
                          'dup':unicode(row['dup'])}]
                    wst_obj.setLayout(wst_obj.getLayout() + l)
                continue
            _id = folder.invokeFactory('WorksheetTemplate', id = 'tmp')
            obj = folder[_id]
            services = row['Service'] and \
                [d.strip() for d in unicode(row['Service']).split(",")] or \
                []
            proxies = services and self.bsc(portal_type="AnalysisService",
                                            getKeyword = services) or []
            services = [p.UID for p in proxies]
            control_ref = self.definitions.get(unicode(row['control_ref']), '')
            blank_ref = self.definitions.get(unicode(row['blank_ref']), '')
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']),
                     Service = services,
                     Layout = [{'pos':unicode(row['pos']),
                                'type':unicode(row['type']),
                                'control_ref':control_ref,
                                'blank_ref':blank_ref,
                                'dup':unicode(row['dup'])}])
            wst_obj = obj
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

    def load_reference_manufacturers(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Reference Manufacturers' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_referencemanufacturers
        for row in rows[3:]:
            row = dict(zip(fields, row))
            _id = folder.invokeFactory('ReferenceManufacturer', id = 'tmp')
            obj = folder[_id]
            obj.edit(title = unicode(row['title']),
                     description = unicode(row['description']))
            obj.unmarkCreationFlag()
            renameAfterCreation(obj)

    def load_partition_setup(self, sheet):
        nr_rows = sheet.get_highest_row()
        nr_cols = sheet.get_highest_column()
##        self.request.response.write("<input type='hidden' id='load_section' value='Reference Manufacturers' max='%s'/>"%(nr_rows-3))
##        self.request.response.flush()
        rows = [[sheet.cell(row=row_nr, column=col_nr).value for col_nr in range(nr_cols)] for row_nr in range(nr_rows)]
        fields = rows[1]
        folder = self.context.bika_setup.bika_analysisservices
        for row in rows[3:]:
            row = dict(zip(fields, row))
            if not row['Analysis Service']: continue
            service = self.services[row['Analysis Service']]
            sampletype = self.sampletypes[row['Sample Type']]
            ps = service.getPartitionSetup()
            containers = []
            if row['Container']:
                for c in row['Container'].split(","):
                    c = c.strip()
                    containers.append(self.containers[c].UID())
            preservations = []
            if row['Preservation']:
                for p in row['Preservation'].split(","):
                    p = p.strip()
                    preservations.append(self.preservations[p].UID())
            ps.append({'sampletype': sampletype.UID(),
                       'container':containers,
                       'preservation':preservations,
                       'separate':row['Separate']})
            service.setPartitionSetup(ps)
