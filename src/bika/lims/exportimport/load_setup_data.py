# -*- coding: utf-8 -*-
#
# This file is part of SENAITE.CORE.
#
# SENAITE.CORE is free software: you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation, version 2.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.
#
# You should have received a copy of the GNU General Public License along with
# this program; if not, write to the Free Software Foundation, Inc., 51
# Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
# Copyright 2018-2020 by it's authors.
# Some rights reserved, see README and LICENSE.

from Products.CMFCore.utils import getToolByName
from bika.lims.browser import BrowserView
from bika.lims import PMF
from bika.lims import logger
from bika.lims.interfaces import ISetupDataImporter
from openpyxl import load_workbook
from pkg_resources import resource_filename
from zope.component import getAdapters
import traceback

import tempfile
import transaction

try:
    from zope.component.hooks import getSite
except:
    # Plone < 4.3
    from zope.app.component.hooks import getSite


class LoadSetupData(BrowserView):

    def __init__(self, context, request):
        BrowserView.__init__(self, context, request)
        self.context = context
        self.request = request

        # dependencies to resolve
        self.deferred = []

        self.request.set('disable_border', 1)

    def solve_deferred(self, deferred=None):
        # walk through self.deferred, linking ReferenceFields as we go
        unsolved = []
        deferred = deferred if deferred else self.deferred
        for d in self.deferred:
            src_obj = d['src_obj']
            src_field = src_obj.getField(d['src_field'])
            multiValued = src_field.multiValued
            src_mutator = src_field.getMutator(src_obj)
            src_accessor = src_field.getAccessor(src_obj)

            tool = getToolByName(self.context, d['dest_catalog'])
            try:
                proxies = tool(d['dest_query'])
            except:
                continue
            if len(proxies) > 0:
                obj = proxies[0].getObject()
                if multiValued:
                    value = src_accessor()
                    value.append(obj.UID())
                else:
                    value = obj.UID()
                src_mutator(value)
            else:
                unsolved.append(d)
        self.deferred = unsolved
        return len(unsolved)

    def __call__(self):
        form = self.request.form
        portal = getSite()
        workbook = None

        if 'setupexisting' in form and 'existing' in form and form['existing']:
                fn = form['existing'].split(":")
                self.dataset_project = fn[0]
                self.dataset_name = fn[1]
                path = 'setupdata/%s/%s.xlsx' % \
                    (self.dataset_name, self.dataset_name)
                filename = resource_filename(self.dataset_project, path)
                try:
                    workbook = load_workbook(filename=filename)  # , use_iterators=True)
                except AttributeError:
                    print ""
                    print traceback.format_exc()
                    print "Error while loading ", path

        elif 'setupfile' in form and 'file' in form and form['file'] and 'projectname' in form and form['projectname']:
                self.dataset_project = form['projectname']
                tmp = tempfile.mktemp()
                file_content = form['file'].read()
                open(tmp, 'wb').write(file_content)
                workbook = load_workbook(filename=tmp)  # , use_iterators=True)
                self.dataset_name = 'uploaded'

        if not workbook:
            message = PMF("File not found...")
            self.context.plone_utils.addPortalMessage(message)
            self.request.RESPONSE.redirect(portal.absolute_url() + "/import")
            return

        adapters = [[name, adapter]
                    for name, adapter
                    in list(getAdapters((self.context, ), ISetupDataImporter))]
        for sheetname in workbook.get_sheet_names():
            transaction.savepoint()
            ad_name = sheetname.replace(" ", "_")
            if ad_name in [a[0] for a in adapters]:
                adapter = [a[1] for a in adapters if a[0] == ad_name][0]
                adapter(self, workbook, self.dataset_project, self.dataset_name)
                adapters = [a for a in adapters if a[0] != ad_name]
        for name, adapter in adapters:
            transaction.savepoint()
            adapter(self, workbook, self.dataset_project, self.dataset_name)

        check = len(self.deferred)
        while len(self.deferred) > 0:
            new = self.solve_deferred()
            logger.info("solved %s of %s deferred references" % (
                check - new, check))
            if new == check:
                raise Exception("%s unsolved deferred references: %s" % (
                    len(self.deferred), self.deferred))
            check = new

        logger.info("Rebuilding bika_setup_catalog")
        bsc = getToolByName(self.context, 'bika_setup_catalog')
        bsc.clearFindAndRebuild()
        logger.info("Rebuilding bika_catalog")
        bc = getToolByName(self.context, 'bika_catalog')
        bc.clearFindAndRebuild()
        logger.info("Rebuilding bika_analysis_catalog")
        bac = getToolByName(self.context, 'bika_analysis_catalog')
        bac.clearFindAndRebuild()

        message = PMF("Changes saved.")
        self.context.plone_utils.addPortalMessage(message)
        self.request.RESPONSE.redirect(portal.absolute_url())
